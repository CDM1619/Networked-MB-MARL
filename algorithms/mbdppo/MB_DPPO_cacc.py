

import time
import matplotlib.pyplot as plt
from numpy.core.numeric import indices
from torch.distributions.normal import Normal
from algorithms.utils import collect, mem_report
from algorithms.models import GaussianActor, GraphConvolutionalModel, MLP, CategoricalActor
from tqdm.std import trange
#from algorithms.algorithm import ReplayBuffer
from ray.state import actors
from gym.spaces.box import Box
from gym.spaces.discrete import Discrete
import torch
import torch.nn as nn
from torch.distributions.categorical import Categorical
from torch.optim import Adam
import numpy as np
import pickle
import copy
import random
from copy import deepcopy as dp
import random
from algorithms.models import CategoricalActor, EnsembledModel, SquashedGaussianActor, ParameterizedModel_MBPPO
import os

test_actors_model = 'xxxxx'   

class MultiCollect:
    def __init__(self, adjacency, device="cuda"):
        """
        Method: 'gather', 'reduce_mean', 'reduce_sum'.
        Adjacency: torch Tensor.
        Everything outward would be in the same device specifed in the initialization parameter.
        """
        self.device = device
        n = adjacency.size()[0]
        adjacency = adjacency > 0 # Adjacency Matrix, with size n_agent*n_agent. 
        adjacency = adjacency | torch.eye(n, device=device).bool() # Should contain self-loop, because an agent should utilize its own info.
        adjacency = adjacency.to(device)
        self.degree = adjacency.sum(dim=1) # Number of information available to the agent.
        self.indices = []
        index_full = torch.arange(n, device=device)
        for i in range(n):
            self.indices.append(torch.masked_select(index_full, adjacency[i])) # Which agents are needed.

    def gather(self, tensor):
        """
        Input shape: [batch_size, n_agent, dim]
        Return shape: [[batch_size, dim_i] for i in range(n_agent)]
        """
        return self._collect('gather', tensor)

    def reduce_mean(self, tensor):
        """
        Input shape: [batch_size, n_agent, dim]
        Return shape: [[batch_size, dim] for i in range(n_agent)]
        """
        return self._collect('reduce_mean', tensor)

    def reduce_sum(self, tensor):
        """
        Input shape: [batch_size, n_agent, dim]
        Return shape: [[batch_size, dim] for i in range(n_agent)]
        """
        return self._collect('reduce_sum', tensor)

    def _collect(self, method, tensor):
        """
        Input shape: [batch_size, n_agent, dim]
        Return shape: 
            gather: [[batch_size, dim_i] for i in range(n_agent)]
            reduce: [batch_size, n_agent, dim]  # same as input
        """
        tensor = tensor.to(self.device)
        if len(tensor.shape) == 1:
            tensor = tensor.unsqueeze(0)
        if len(tensor.shape) == 2:
            tensor = tensor.unsqueeze(-1)
        b, n, depth = tensor.shape
        result = []
        for i in range(n):
            if method == 'gather':
                result.append(torch.index_select(tensor, dim=1, index=self.indices[i]).view(b, -1))
            elif method == 'reduce_mean':
                result.append(torch.index_select(tensor, dim=1, index=self.indices[i]).mean(dim=1))
            else:
                result.append(torch.index_select(tensor, dim=1, index=self.indices[i]).sum(dim=1))
        if method != 'gather':
            result = torch.stack(result, dim=1)
        return result

class Trajectory:
    def __init__(self, **kwargs):
        """
        Data are of size [T, n_agent, dim].
        """
        self.names = ["s", "a", "r", "s1", "d", "logp"]
        self.dict = {name: kwargs[name] for name in self.names}
        self.length = self.dict["s"].size()[0]
    
    def getFraction(self, length, start=None):
        if self.length < length:
            length = self.length
        start_max = self.length - length
        if start is None:
            start = torch.randint(low=0, high=start_max+1, size=(1,)).item()
        if start > start_max:
            start = start_max
        if start < 0:
            start = 0
        new_dict = {name: self.dict[name][start:start+length] for name in self.names}
        return Trajectory(**new_dict)
    
    def __getitem__(self, key):
        assert key in self.names
        return self.dict[key]
    
    @classmethod
    def names(cls):
        return ["s", "a", "r", "s1", "d", "logp"]

class TrajectoryBuffer:
    def __init__(self, device="cuda"):
        self.device = device
        self.s, self.a, self.r, self.s1, self.d, self.logp = [], [], [], [], [], []
    
    def store(self, s, a, r, s1, d, logp):
        """
        Would be converted into [batch_size, n_agent, dim].
        """
        device = self.device
        [s, r, s1, logp] = [torch.as_tensor(item, device=device, dtype=torch.float) for item in [s, r, s1, logp]]
        d = torch.as_tensor(d, device=device, dtype=torch.bool)
        a = torch.as_tensor(a, device=device)
        while s.dim() <= 2:
            s = s.unsqueeze(dim=0)
        b, n, dim = s.size()
        if d.dim() <= 1:
            d = d.unsqueeze(0)
        d = d[:, :n]
        if r.dim() <= 1:
            r = r.unsqueeze(0)
        r = r[:, :n]
        [s, a, r, s1, d, logp] = [item.view(b, n, -1) for item in [s, a, r, s1, d, logp]]
        self.s.append(s)
        self.a.append(a)
        self.r.append(r)
        self.s1.append(s1)
        self.d.append(d)
        self.logp.append(logp)
    
    def retrieve(self, length=None):
        """
        Returns trajectories with s, a, r, s1, d, logp.
        Data are of size [T, n_agent, dim]
        """
        names = ["s", "a", "r", "s1", "d", "logp"]
        trajs = []
        traj_all = {}
        if self.s == []:
            return []
        for name in names:
            traj_all[name] = torch.stack(self.__getattribute__(name), dim=1)
        n = traj_all['s'].size()[0]
        for i in range(n):
            traj_dict = {}
            for name in names:
                traj_dict[name] = traj_all[name][i]  #ndecth batch into single traj
            trajs.append(Trajectory(**traj_dict))
        return trajs

class ModelBuffer:
    def __init__(self, max_traj_num):
        self.max_traj_num = max_traj_num
        self.trajectories = []
        self.ptr = -1
        self.count = 0
    
    def storeTraj(self, traj):
        if self.count < self.max_traj_num:
            self.trajectories.append(traj)
            self.ptr = (self.ptr + 1) % self.max_traj_num
            self.count = min(self.count + 1, self.max_traj_num)
        else:
            self.trajectories[self.ptr] = traj
            self.ptr = (self.ptr + 1) % self.max_traj_num
    
    def storeTrajs(self, trajs):
        for traj in trajs:
            self.storeTraj(traj)
    
    def sampleTrajs(self, n_traj):
        traj_idxs = np.random.choice(range(self.count), size=(n_traj,), replace=True)
        return [self.trajectories[i] for i in traj_idxs]

class OnPolicyRunner:
    def __init__(self, logger, run_args, alg_args, agent, env_learn, env_test,env_args, **kwargs):
        self.logger = logger
        self.name = run_args.name
        if not run_args.init_checkpoint is None:
            agent.load(run_args.init_checkpoint)
            logger.log(interaction=run_args.start_step)  
            
        # agent.load_state_dict(torch.load('C:/Users/86153/Desktop/MB-MARL/checkpoints/standard_UAV_Env_DPPOAgent_9676/1618304_-130.85930694345433.pt'))
        # agent.load_state_dict(torch.load('./123.pt'))  
        # load_state_dict(torch.load('./Models/49999best_actor.pt'))
        
        self.env_name = env_args.env        
        self.start_step = run_args.start_step 
        # algorithm arguments
        self.n_iter = alg_args.n_iter
        self.n_inner_iter = alg_args.n_inner_iter
        self.n_warmup = alg_args.n_warmup
        self.n_model_update = alg_args.n_model_update
        self.n_model_update_warmup = alg_args.n_model_update_warmup
        self.n_test = alg_args.n_test
        self.test_interval = alg_args.test_interval
        self.rollout_length = alg_args.rollout_length
        
        self.test_length = alg_args.test_length
        # self.test_length = 100
        
        self.max_episode_len = alg_args.max_episode_len
        self.clip_scheme = None if (not hasattr(alg_args, "clip_scheme")) else alg_args.clip_scheme
        
        # agent initialization
        self.agent = agent
        self.device = self.agent.device if hasattr(self.agent, "device") else "cpu"

        # environment initialization
        self.env_learn = env_learn
        self.env_test = env_test
        
        self.n_agent = env_test.n_agent
        # self.n_agent = 28
        
        # buffer initialization
        self.discrete = agent.discrete
        action_dtype = torch.long if self.discrete else torch.float
        self.model_based = alg_args.model_based
        self.model_batch_size = alg_args.model_batch_size
        if self.model_based:
            self.n_traj = alg_args.n_traj
            self.model_traj_length = alg_args.model_traj_length
            self.model_error_thres = alg_args.model_error_thres
            self.model_buffer = ModelBuffer(alg_args.model_buffer_size)
            self.model_update_length = alg_args.model_update_length
            self.model_validate_interval = alg_args.model_validate_interval
            self.model_length_schedule = alg_args.model_length_schedule
            self.model_prob = alg_args.model_prob
        self.s, self.episode_len, self.episode_reward = self.env_learn.reset(), 0, 0

        # load pretrained model
        self.load_pretrained_model = alg_args.load_pretrained_model
        if self.model_based and self.load_pretrained_model:
            self.agent.load_model(alg_args.pretrained_model)

    def run(self):
        if self.model_based and not self.load_pretrained_model:
            for _ in trange(self.n_warmup):
                trajs = self.rollout_env()
                self.model_buffer.storeTrajs(trajs)
            self.updateModel(self.n_model_update_warmup) # Sample trajectories, then shorten them.
        for iter in trange(self.n_iter):
            if iter % self.test_interval == 0:
                mean_return = self.test()
                self.agent.save(info = mean_return)
            trajs = self.rollout_env()  #  TO cheak: rollout n_step, maybe multi trajs
            if self.model_based:
                self.model_buffer.storeTrajs(trajs)
                self.updateModel()
            agentInfo = []
            real_trajs = trajs
            for inner in trange(self.n_inner_iter):
                if self.model_based:
                    use_model = np.random.uniform() < self.model_prob
                    if use_model:
                        if self.model_length_schedule is not None:
                            trajs = self.rollout_model(real_trajs, self.model_length_schedule(iter))
                        else:
                            trajs = self.rollout_model(real_trajs)
                    else:
                        trajs = trajs
                if self.clip_scheme is not None:
                    info = self.agent.updateAgent(trajs, self.clip_scheme(iter))     #  TO cheak: updata
                else:
                    info = self.agent.updateAgent(trajs)
                agentInfo.append(info)
                if self.agent.checkConverged(agentInfo):
                    break
            self.logger.log(inner_iter = inner + 1, iter=iter)

    def test(self):
        """
        The environment should return sth like [n_agent, dim] or [batch_size, n_agent, dim] in either numpy or torch.
        """
        time_t = time.time()
        length = self.test_length
        returns = []
        scaled = []
        lengths = []
        episodes = []

        for i in trange(self.n_test):
           
# test cacc---------------------------------------------------------------------------------------------------------

            v = [[]]*self.n_agent 
            v_d = [[]]*self.n_agent 
            vh_d = [[]]*self.n_agent 
            h = [[]]*self.n_agent 
            u = [[]]*self.n_agent 
            Step = []


            v_ = []
            v_d_ = []
            vh_d_ = []
            h_ = []
            u_ = []

            
# --------------------------------------------------------------------------------------------------------------

   
            episode = []
            env = self.env_test
            env.reset()
            d, ep_ret, ep_len = np.array([False]), 0, 0
            # while not(d.any() or (ep_len == length)):
            while not (ep_len == length):
                Step.append(ep_len)
                s = env.get_state_() # dim = 2 or 3 (vectorized)
                s = torch.as_tensor(s, dtype=torch.float, device=self.device)
                a = self.agent.act(s).sample() # a is a tensor
                #a = self.agent.act(s)
                # a = a.tanh()
                
                #print("s=",s)

                a = a.detach().cpu().numpy() # might not be squeezed at the last dimension. env should deal with this though.
                
                
                
                #让第一辆车随机行驶
                print("action=",a)           
                a1 = random.randint(0,3)
                a[0]=a1
                
                
                s1, r, d, _ = env.step(a)
                episode += [(s.tolist(), a.tolist(), r.tolist())]
                d = np.array(d)
                ep_ret += r.sum()
                ep_len += 1
                self.logger.log(interaction=None)   













                
                s1 = dp(s1)
                # s1 = dp(env.get_state_real(dp(s1)))
                # a = dp(env.get_action_real(dp(a)))               
                for k in range(self.n_agent):
                    s1 = dp(s1)
                    # a = dp(a)
                    k = dp(k)
                    v[k].append(s1[k][0]*15 + 15)
                    h[k].append(s1[k][3]*20 + 20)

                
                # episode += [(s.tolist(), a.tolist(), r.tolist())]
                # d = np.array(d)
                # ep_ret += r.sum()
                # ep_len += 1
                # self.logger.log(interaction=None)   

            
# --------------------------------------------------------------------------------------------------------------   

# test uav : plot_trajectory----------------------------------------------------------------------------------------------------   

            # print('sss=',len(p_x))
            # print('sss=',len(X[0]))
            # print('a=',len(X))
            # print('bbb=',X)
            color = ['red','royalblue','deeppink','lightpink','gold','green','peru','fuchsia','teal','darkred','orange','orange']
            p_v=dp(v[0])
            p_h=dp(h[0])
            # print('sss=',len(p_x))           
          
            for j in range(self.n_agent):
                temp_v=[]
                temp_h=[]
                
                q=dp(j)
                for r in range(self.test_length):
                    r=dp(r)
                    temp_v.append(dp(p_v[q+r*self.n_agent]))
                    temp_h.append(dp(p_h[q+r*self.n_agent]))

                v_.append(dp(temp_v))
                h_.append(dp(temp_h))
# ?????2ì?1ì?￡------------------------------------------------------------------------------------------------------    
            import matplotlib as mpl
            import matplotlib.pyplot as plt
            from matplotlib.font_manager import FontProperties
            #mpl.rcParams['font.sans-serif'] = 'Tahoma' 
            if i+1 == self.n_test:                
                #fig = plt.figure(1)
                mpl.rcParams['font.sans-serif'] = 'Arial' 
                plt.figure(figsize=(8,6))
                mpl.style.use('seaborn')
                plt.grid(color='white',linewidth=3)
                for j in range(self.n_agent):  

                    #if j == 0:
                        #plt.plot(Step,dp(v_)[j],color='red',label="vehicle 1")
                    if j == 1:
                        plt.plot(Step,dp(v_)[j],color='fuchsia',label="vehicle 2")
                    if j == 2:
                        plt.plot(Step,dp(v_)[j],color='green',label="vehicle 3")
                    if j == 3:
                        plt.plot(Step,dp(v_)[j],color='orange',label="vehicle 4")
                    #if j == 4:
                        #plt.plot(Step,dp(v_)[j],color='royalblue',label="vehicle 5")
                    #if j == 5:
                        #plt.plot(Step,dp(v_)[j],color='teal',label="vehicle 6")
                    #if j == 6:
                        #plt.plot(Step,dp(v_)[j],color='peru',label="vehicle 7")
                    #if j == 7:
                        #plt.plot(Step,dp(v_)[j],color='black',label='vehicle 8')
                        

                plt.xlabel('Steps',fontsize=15)
                plt.ylabel('Vehicle velocity(m/s)',fontsize=15)
                plt.legend(fontsize=15,loc='lower right')
                plt.xticks(size = 15)
                plt.yticks(size = 15)
                plt.legend(fontsize=15,loc='upper right')
                #plt.ylim(10,30)
                #plt.grid()  
                
                # plt.savefig('./catchup-velocity.png',bbox_inches = 'tight',dpi=1200)
                # plt.savefig('./catchup-velocity.svg',bbox_inches = 'tight',dpi=1200)
                # plt.savefig('./catchup-velocity.pdf',bbox_inches = 'tight',dpi=1200)     

                plt.savefig('./slowdonw-velocity.png',bbox_inches = 'tight',dpi=1200)
                plt.savefig('./slowdown-velocity.svg',bbox_inches = 'tight',dpi=1200)
                plt.savefig('./slowdown-velocity.pdf',bbox_inches = 'tight',dpi=1200)   
                       
                plt.show()

            if i+1 == self.n_test:                
                #fig = plt.figure(2)
                mpl.rcParams['font.sans-serif'] = 'Arial' 
                plt.figure(figsize=(8,6))
                mpl.style.use('seaborn')
                plt.grid(color='white',linewidth=3)
                for j in range(self.n_agent):  
                        
                    #if j == 0:
                        #plt.plot(Step,dp(h_)[j],color='red',label="vehicle 1")
                    if j == 1:
                        plt.plot(Step,dp(h_)[j],color='fuchsia',label="vehicle 2")
                    if j == 2:
                        plt.plot(Step,dp(h_)[j],color='green',label="vehicle 3")
                    if j == 3:
                        plt.plot(Step,dp(h_)[j],color='orange',label="vehicle 4")
                    #if j == 4:
                        #plt.plot(Step,dp(h_)[j],color='royalblue',label="vehicle 5")
                    #if j == 5:
                        #plt.plot(Step,dp(h_)[j],color='teal',label="vehicle 6")
                    #if j == 6:
                        #plt.plot(Step,dp(h_)[j],color='peru',label="vehicle 7")
                    #if j == 7:
                        #plt.plot(Step,dp(h_)[j],color='black',label='vehicle 8')
                plt.xlabel('Steps',fontsize=15)
                plt.ylabel('Vehicle headway(m)',fontsize=15)
                plt.legend(fontsize=15,loc='lower right')
                plt.xticks(size = 15)
                plt.yticks(size = 15)
                plt.legend(fontsize=15,loc='upper right')
                #plt.ylim(15,30)
                #plt.grid()
                # plt.savefig('./catchup-headway.png',bbox_inches = 'tight',dpi=1200)
                # plt.savefig('./catchup-headway.svg',bbox_inches = 'tight',dpi=1200)
                # plt.savefig('./catchup-headway.pdf',bbox_inches = 'tight',dpi=1200) 
                
                plt.savefig('./slowdown-headway.png',bbox_inches = 'tight',dpi=1200)
                plt.savefig('./slowdown-headway.svg',bbox_inches = 'tight',dpi=1200)
                plt.savefig('./slowdown-headway.pdf',bbox_inches = 'tight',dpi=1200)  
                            
                plt.show()


# --------------------------------------------------------------------------------------------------------------   
            if i+1 == self.n_test:  
                from openpyxl import Workbook              
                for j in range(self.n_agent):  

                    if j == 1:
                        
                        from openpyxl import Workbook
                        wb_1 = Workbook() #′′?¨1¤×÷2?
                        ws_1 = wb_1.active #?¤??1¤×÷±í
                        ws_1['A1'] = 'Step'
                        ws_1['B1'] = 'v'                            
                        for i in range(len(Step)):                
                            ws_1.append([Step[i],dp(v_)[j][i]])              
                        wb_1.save('./v1.xlsx') 


                        wb_1 = Workbook() #′′?¨1¤×÷2?
                        ws_1 = wb_1.active #?¤??1¤×÷±í
                        ws_1['A1'] = 'Step'
                        ws_1['B1'] = 'pos'                            
                        for i in range(len(Step)):                
                            ws_1.append([Step[i],dp(h_)[j][i]])              
                        wb_1.save('./p1.xlsx') 
              
                            
                    if j == 2:

                        from openpyxl import Workbook
                        wb_2 = Workbook() #′′?¨1¤×÷2?
                        ws_2 = wb_2.active #?¤??1¤×÷±í
                        ws_2['A1'] = 'Step'
                        ws_2['B1'] = 'v'                            
                        for i in range(len(Step)):                
                            ws_2.append([Step[i],dp(v_)[j][i]])              
                        wb_2.save('./v2.xlsx')  

                        wb_2 = Workbook() #′′?¨1¤×÷2?
                        ws_2 = wb_2.active #?¤??1¤×÷±í
                        ws_2['A1'] = 'Step'
                        ws_2['B1'] = 'pos'                            
                        for i in range(len(Step)):                
                            ws_2.append([Step[i],dp(h_)[j][i]])              
                        wb_2.save('./p2.xlsx')  

                    if j == 3:

                        from openpyxl import Workbook
                        wb_3 = Workbook() #′′?¨1¤×÷2?
                        ws_3 = wb_3.active #?¤??1¤×÷±í
                        ws_3['A1'] = 'Step'
                        ws_3['B1'] = 'v'                            
                        for i in range(len(Step)):                
                            ws_3.append([Step[i],dp(v_)[j][i]])              
                        wb_3.save('./v3.xlsx')  

                        wb_3 = Workbook() #′′?¨1¤×÷2?
                        ws_3 = wb_3.active #?¤??1¤×÷±í
                        ws_3['A1'] = 'Step'
                        ws_3['B1'] = 'pos'                            
                        for i in range(len(Step)):                
                            ws_3.append([Step[i],dp(h_)[j][i]])              
                        wb_3.save('./p3.xlsx') 







                
            from openpyxl import Workbook
       
            if i+1 == self.n_test:                 
                
                data = env.traffic_data
                time_sec = []
                avg_queue = []
                avg_wait_sec = []
                for d in range(len(data)):
                    #time_sec.append(data[d]['time_sec'])
                    #avg_queue.append(data[d]['avg_headway_m'])
                    #avg_wait_sec.append(data[d]['avg_speed_mps'])
                    print('type=',len(data[d]['avg_headway_m']))
                    time_sec.append(data[d]['time_sec'][1])
                    avg_queue.append(data[d]['avg_headway_m'][1])
                    avg_wait_sec.append(data[d]['avg_speed_mps'][1])


    
    
                wb_1 = Workbook() #′′?¨1¤×÷2?
                ws_1 = wb_1.active #?¤??1¤×÷±í
                ws_1['A1'] = 'time_sec'
                ws_1['B1'] = 'avg_queue'                            
                for ii in range(len(time_sec)):                
                    ws_1.append([time_sec[ii],avg_queue[ii]])              
                wb_1.save('./avg_headway_m'+'.xlsx') 


                wb_2 = Workbook() #′′?¨1¤×÷2?
                ws_2 = wb_2.active #?¤??1¤×÷±í
                ws_2['A1'] = 'time_sec'
                ws_2['B1'] = 'avg_wait_sec'                            
                for ii in range(len(time_sec)):                
                    ws_2.append([time_sec[ii],avg_wait_sec[ii]])              
                wb_2.save('./avg_speed_mps'+'.xlsx') 

                





#-------------------------------------------------------------------------------------------------



#-------------------------------------------------------------------------------------------------



                
# # # -------------------------------------------------------------------------------------------------------------  
            

#             if hasattr(env, 'rescaleReward'):
#                 scaled += [ep_ret]
#                 ep_ret = env.rescaleReward(ep_ret, ep_len)
#             returns += [ep_ret]
#             lengths += [ep_len]
#             episodes += [episode]
#         returns = np.stack(returns, axis=0)
#         lengths = np.stack(lengths, axis=0)
#         self.logger.log(test_episode_reward=returns, test_episode_len=lengths, test_round=None)
#         print(returns)
#         print(f"{self.n_test} episodes average accumulated reward: {returns.mean()}")
#         if hasattr(env, 'rescaleReward'):
#             print(f"scaled reward {np.mean(scaled)}")
#         with open(f"checkpoints/{self.name}/test.pickle", "wb") as f:
#             pickle.dump(episodes, f)
#         with open(f"checkpoints/{self.name}/test.txt", "w") as f:
#             for episode in episodes:
#                 for step in episode:
#                     f.write(f"{step[0]}, {step[1]}, {step[2]}\n")
#                 f.write("\n")
#         self.logger.log(test_time=time.time()-time_t)
#         return returns.mean()
    





#     def test(self):
#         """
#         The environment should return sth like [n_agent, dim] or [batch_size, n_agent, dim] in either numpy or torch.
#         """
#         time_t = time.time()
#         length = self.test_length
#         returns = []
#         scaled = []
#         lengths = []
#         episodes = []
#         for i in trange(self.n_test):
           
# # test uav---------------------------------------------------------------------------------------------------------

#             X = [[]]*self.n_agent 
#             Y = [[]]*self.n_agent 
#             # Z = [[]]*self.n_agent 
#             # Pusin = [[]]*self.n_agent 
#             # Gamma = [[]]*self.n_agent 
#             # Mju = [[]]*self.n_agent 


#             X_ = []
#             Y_ = []
#             # Z_ = []
#             # Pusin_ = []
#             # Gamma_ = []
#             # Mju_ = []
            
# # --------------------------------------------------------------------------------------------------------------

   
#             episode = []
#             env = self.env_test
#             env.reset()
#             d, ep_ret, ep_len = np.array([False]), 0, 0
#             # while not(d.any() or (ep_len == length)):
#             while not (ep_len == length):
#                 s = env.get_state_() # dim = 2 or 3 (vectorized)
#                 s = torch.as_tensor(s, dtype=torch.float, device=self.device)
#                 a = self.agent.act(s).sample() # a is a tensor
#                 # a = self.agent.act(s)
#                 a = a.tanh()

#                 a = a.detach().cpu().numpy() # might not be squeezed at the last dimension. env should deal with this though.
#                 s1, r, d, _ = env.step(a)
#                 episode += [(s.tolist(), a.tolist(), r.tolist())]
#                 d = np.array(d)
#                 ep_ret += r.sum()
#                 ep_len += 1
#                 self.logger.log(interaction=None)   

#                 s1 = dp(env.get_state_real(dp(s1)))
#                 a = dp(env.get_action_real(dp(a))) 
                
#                 for k in range(self.n_agent):
#                     s1 = dp(s1)
#                     a = dp(a)
#                     k = dp(k)
#                     X[k].append(s1[k][0])
#                     Y[k].append(s1[k][1])
#                     # Z[k].append(s1[k][2])
#                     # Pusin[k].append(s1[k][3])
#                     # Gamma[k].append(s1[k][4])
#                     # Mju[k].append(s1[k][6])  

                
#                 # episode += [(s.tolist(), a.tolist(), r.tolist())]
#                 # d = np.array(d)
#                 # ep_ret += r.sum()
#                 # ep_len += 1
#                 # self.logger.log(interaction=None)   
            
# # --------------------------------------------------------------------------------------------------------------   

# # test uav : plot_trajectory----------------------------------------------------------------------------------------------------   

#             color = ['red','royalblue','deeppink','lightpink','gold','green','peru','fuchsia','teal','darkred','orange','orange']
#             p_x=dp(X[0])
#             p_y=dp(Y[0])
#             # p_z=dp(Z[0])
#             # p_pusin=dp(Pusin[0])
#             # p_gamma=dp(Gamma[0])
#             # p_mju=dp(Mju[0])
#             # print('sss=',len(p_x))           
          
#             for j in range(self.n_agent):
#                 temp_x=[]
#                 temp_y=[]
#                 # temp_z=[]
#                 # temp_pusin=[]
#                 # temp_gamma=[]
#                 # temp_mju=[]
                
#                 q=dp(j)
#                 for r in range(self.test_length):
#                     r=dp(r)
#                     temp_x.append(dp(p_x[q+r*self.n_agent]))
#                     temp_y.append(dp(p_y[q+r*self.n_agent]))
#                     # temp_z.append(dp(p_z[q+r*self.n_agent]))
#                     # temp_pusin.append(dp(p_pusin[q+r*self.n_agent]))
#                     # temp_gamma.append(dp(p_gamma[q+r*self.n_agent]))
#                     # temp_mju.append((p_mju[q+r*self.n_agent]))
#                 X_.append(dp(temp_x))
#                 Y_.append(dp(temp_y))
#                 # Z_.append(dp(temp_z))
#                 # Pusin_.append(dp(temp_pusin))
#                 # Gamma_.append(dp(temp_gamma))
#                 # Mju_.append(dp(temp_mju))

# # ?????2ì?1ì?￡------------------------------------------------------------------------------------------------------    


#             # print('sss=',X_[0])
#             if i+1 == self.n_test:
#                 fig = plt.figure(1)

#                 for j in range(self.n_agent):  

#                     c=random.randint(0,len(color)-1)
                    
#                     plt.plot(dp(X_[j]), dp(Y_[j]), color[c], alpha = 0.3,ms=1, linestyle = '-')

#                 plt.savefig('./Trajectory.png',bbox_inches = 'tight')
#                 plt.show()

# #-------------------------------------------------------------------------------------------------



#             if hasattr(env, 'rescaleReward'):
#                 scaled += [ep_ret]
#                 ep_ret = env.rescaleReward(ep_ret, ep_len)
#             returns += [ep_ret]
#             lengths += [ep_len]
#             episodes += [episode]
#         returns = np.stack(returns, axis=0)
#         lengths = np.stack(lengths, axis=0)
#         self.logger.log(test_episode_reward=returns, test_episode_len=lengths, test_round=None)
#         print(returns)
#         print(f"{self.n_test} episodes average accumulated reward: {returns.mean()}")
#         if hasattr(env, 'rescaleReward'):
#             print(f"scaled reward {np.mean(scaled)}")
#         with open(f"checkpoints/{self.name}/test.pickle", "wb") as f:
#             pickle.dump(episodes, f)
#         with open(f"checkpoints/{self.name}/test.txt", "w") as f:
#             for episode in episodes:
#                 for step in episode:
#                     f.write(f"{step[0]}, {step[1]}, {step[2]}\n")
#                 f.write("\n")
#         self.logger.log(test_time=time.time()-time_t)
#         return returns.mean()







    

    def rollout_env(self, length = 0):
        """
        The environment should return sth like [n_agent, dim] or [batch_size, n_agent, dim] in either numpy or torch.
        """
        time_t = time.time()
        if length <= 0:
            length = self.rollout_length
        env = self.env_learn
        trajs = []
        traj = TrajectoryBuffer(device=self.device)
        
        start = time.time()
        for t in range(length):
            s = env.get_state_()
            s = torch.as_tensor(s, dtype=torch.float, device=self.device)
            dist = self.agent.act(s)
            a = dist.sample()
            logp = dist.log_prob(a)
            # a = a.tanh()
            a = a.detach().cpu().numpy()
            s1, r, d, _ = env.step(a)
            traj.store(s, a, r, s1, d, logp)
            episode_r = r
            if hasattr(env, '_comparable_reward'):
                episode_r = env._comparable_reward()
            if episode_r.ndim > 1:
                episode_r = episode_r.mean(axis=0)
            self.episode_reward += episode_r
            self.episode_len += 1
            self.logger.log(interaction=None)
            if self.episode_len == self.max_episode_len:
                d = np.zeros(d.shape, dtype=np.float32)
            d = np.array(d)
            #if d.any() or (self.episode_len == self.max_episode_len):
            if self.episode_len == self.max_episode_len:
                self.logger.log(episode_reward=self.episode_reward.sum(), episode_len = self.episode_len, episode=None)
                try:
                    _, self.episode_reward, self.episode_len = self.env_learn.reset(), 0, 0#TODO:catch up the error
                except Exception as e:
                    print('reset error!:', e)
                    _, self.episode_reward, self.episode_len = self.env_learn.reset(), 0, 0  # TODO:catch up the error
                trajs += traj.retrieve()
                traj = TrajectoryBuffer(device=self.device)
        end = time.time()
        print('time in 1 episode is ',end-start)
        trajs += traj.retrieve(length=self.max_episode_len)
        self.logger.log(env_rollout_time=time.time()-time_t)
        return trajs
    
    def rollout_model(self, trajs, length=0):
        time_t = time.time()
        n_traj = self.n_traj
        if length <= 0:
            length = self.model_traj_length
        s = [traj['s'] for traj in trajs]
        s = torch.stack(s, dim=0)
        b, T, n, depth = s.shape
        s = s.view(-1, n, depth)
        idxs = torch.randint(low=0, high=b * T, size=(n_traj,), device=self.device)
        s = s.index_select(dim=0, index=idxs)
        # s.dim() == 3

        trajs = TrajectoryBuffer(device=self.device)
        for _ in range(length):
            #a, logp = self.agent.act(s, requires_log=True)
            dist = self.agent.act(s)
            a = dist.sample()
            logp = dist.log_prob(a)
            r, s1, d, _ = self.agent.model_step(s, a)
            trajs.store(s, a, r, s1, d, logp)
            s = s1
        trajs = trajs.retrieve()
        self.logger.log(model_rollout_time=time.time()-time_t)
        return trajs
    
    def updateModel(self, n=0):
        if n <= 0:
            n = self.n_model_update
        for i_model_update in trange(n):
            trajs = self.model_buffer.sampleTrajs(self.model_batch_size)
            trajs = [traj.getFraction(length=self.model_update_length) for traj in trajs]
            self.agent.updateModel(trajs, length=self.model_update_length)
            if i_model_update % self.model_validate_interval == 0:
                validate_trajs = self.model_buffer.sampleTrajs(self.model_batch_size)
                validate_trajs = [traj.getFraction(length=self.model_update_length) for traj in validate_trajs]
                rel_error = self.agent.validateModel(validate_trajs, length=self.model_update_length)
                if rel_error < self.model_error_thres:
                    break
        self.logger.log(model_update = i_model_update + 1)
    
    def testModel(self, n = 0):
        trajs = self.model_buffer.sampleTrajs(self.model_batch_size)
        trajs = [traj.getFraction(length=self.model_update_length) for traj in trajs]
        return self.agent.validateModel(trajs, length=self.model_update_length)

class IA2C(nn.ModuleList):
    def __init__(self, logger, device, agent_args, **kwargs):
        super().__init__()
        self.logger = logger
        self.device = device
        self.n_agent = agent_args.n_agent
        self.gamma = agent_args.gamma
        self.lamda = agent_args.lamda
        self.clip = agent_args.clip
        self.target_kl = agent_args.target_kl
        self.v_coeff = agent_args.v_coeff
        self.v_thres = agent_args.v_thres
        self.entropy_coeff = agent_args.entropy_coeff
        self.entropy_coeff_decay = agent_args.entropy_coeff_decay  # only in IA2C
        self.lr = agent_args.lr
        self.lr_v = agent_args.lr_v
        self.n_update_v = agent_args.n_update_v
        self.n_update_pi = agent_args.n_update_pi
        self.n_minibatch = agent_args.n_minibatch
        self.use_reduced_v = agent_args.use_reduced_v
        self.use_rtg = agent_args.use_rtg
        self.use_gae_returns = agent_args.use_gae_returns

        self.advantage_norm = agent_args.advantage_norm

        self.observation_dim = agent_args.observation_dim
        self.action_space = agent_args.action_space
        self.discrete = isinstance(agent_args.action_space, Discrete)
        if self.discrete:
            self.action_dim = self.action_space.n
            self.action_shape = self.action_dim
        else:
            self.action_shape = self.action_space.shape
            self.action_dim = 1
            for j in self.action_shape:
                self.action_dim *= j
            self.action_low = self.action_space.low.item()
            self.action_high = self.action_space.high.item()
            self.squeeze = agent_args.squeeze
        # if adj diag is not one, we should add a eye matrix
        agent_args.adj = (torch.as_tensor(agent_args.adj, device=self.device, dtype=torch.float)>0) | torch.eye(self.n_agent, device=device).bool()
        self.adj = torch.as_tensor(agent_args.adj, device=self.device, dtype=torch.float)
        self.radius_v = agent_args.radius_v
        self.radius_pi = agent_args.radius_pi
        self.pi_args = agent_args.pi_args
        self.v_args = agent_args.v_args
        self.collect_pi, self.actors = self._init_actors()
        self.collect_v, self.vs = self._init_vs()

        self.optimizer_v = Adam(self.vs.parameters(), lr=self.lr_v)
        self.optimizer_pi = Adam(self.actors.parameters(), lr=self.lr)

    def get_logp(self, s, a):
        """
        Requires input of [batch_size, n_agent, dim] or [n_agent, dim].
        Returns a tensor whose dim() == 3.
        """
        s = torch.as_tensor(s, dtype=torch.float32, device=self.device)
        # print('dim is',s)
        dim = s.dim()
        # print('dim is',dim)
        while s.dim() <= 2:
            s = s.unsqueeze(0)
            a = a.unsqueeze(0)
        while a.dim() < s.dim():
            a = a.unsqueeze(-1)
        s = self.collect_pi.gather(s)
        # Now s[i].dim() == 2, a.dim() == 3
        log_prob = []
        for i in range(self.n_agent):
            if self.discrete:
                probs = self.actors[i](s[i])
                log_prob.append(torch.log(torch.gather(probs, dim=-1, index=torch.select(a, dim=1, index=i).long())))
            else:
                log_prob.append(self.actors[i](s[i], a.select(dim=1, index=i)))
        log_prob = torch.stack(log_prob, dim=1)
        while log_prob.dim() < 3:
            log_prob = log_prob.unsqueeze(-1)
        return log_prob

    def act(self, s, requires_log=False):
        """
        Requires input of [batch_size, n_agent, dim] or [n_agent, dim].
        This method is gradient-free. To get the gradient-enabled probability information, use get_logp().
        Returns a distribution with the same dimensions of input.
        """
        with torch.no_grad():
            dim = s.dim()
            while s.dim() <= 2:
                s = s.unsqueeze(0)
            s = s.to(self.device)
            s = self.collect_pi.gather(s)  # all state into [ self +  ]
            # Now s[i].dim() == 2 ([batch_size, dim])

            if self.discrete:
                probs = []
                for i in range(self.n_agent):
                    probs.append(self.actors[i](s[i]))
                probs = torch.stack(probs, dim=1)
                return Categorical(probs)
            else:
                means, stds = [], []
                for i in range(self.n_agent):
                    mean, std = self.actors[i](s[i])
                    means.append(mean)
                    stds.append(std)
                means = torch.stack(means, dim=1)
                stds = torch.stack(stds, dim=1)
                while means.dim() > dim:
                    means = means.squeeze(0)
                    stds = stds.squeeze(0)
                return Normal(means, stds)


    def _process_traj(self, s, a, r, s1, d, logp):
        """
        Input are all in shape [batch_size, T, n_agent, dim]
        """
        with torch.no_grad():
            b, T, n, dim_s = s.shape
            s, a, r, s1, d, logp = [item.to(self.device) for item in [s, a, r, s1, d, logp]]
            value = self._evalV(s.view(-1, n, dim_s)).view(b, T, n, -1)

            returns = torch.zeros(value.size(), device=self.device)
            deltas, advantages = torch.zeros_like(returns), torch.zeros_like(returns)
            prev_value = self._evalV(s1.select(1, T - 1))
            if not self.use_rtg:
                prev_return = prev_value
            else:
                prev_return = torch.zeros_like(prev_value)
            prev_advantage = torch.zeros_like(prev_return)
            d_mask = d.float()
            for t in reversed(range(T)):
                deltas[:, t, :, :]= r.select(1, t) + self.gamma * (1-d_mask.select(1,t)) * prev_value - value.select(1, t).detach()
                advantages[:, t, :, :] = deltas.select(1, t) + self.gamma * self.lamda * (1-d_mask.select(1,t)) * prev_advantage
                if self.use_gae_returns:
                    returns[:, t, :, :] = value.select(1, t).detach() + advantages.select(1, t)
                else:
                    returns[:, t, :, :] = r.select(1, t) + self.gamma * (1-d_mask.select(1, t)) * prev_return

                prev_return = returns.select(1, t)
                prev_value = value.select(1, t)
                prev_advantage = advantages.select(1, t)
            reduced_advantages = self.collect_v.reduce_sum(advantages.view(-1, n, 1)).view(advantages.size())
            if self.advantage_norm and reduced_advantages.size()[1] > 1:
                reduced_advantages = (reduced_advantages - reduced_advantages.mean(dim=1, keepdim=True)) / (reduced_advantages.std(dim=1, keepdim=True) + 1e-5)
                advantages = (advantages - advantages.mean(dim=1, keepdim=True)) / (advantages.std(dim=1, keepdim=True) + 1e-5)
        return value, returns, advantages, reduced_advantages

    def load(self):
        # set  run_args.init_checkpoint  = None
        pass

    def checkConverged(self, ls_info):
        #TODO: not neccessary
        return False

    def save(self, info=None):
        self.logger.save(self, info=info)

    def _evalV(self, s):
        # Requires input in shape [-1, n_agent, dim]
        s = s.to(self.device)
        s = self.collect_v.gather(s)
        values = []
        for i in range(self.n_agent):
            values.append(self.vs[i](s[i]))
        return torch.stack(values, dim=1)

    def _init_actors(self):
        collect_pi = MultiCollect(torch.matrix_power(self.adj, self.radius_pi), device=self.device)
        actors = nn.ModuleList()
        for i in range(self.n_agent):
            self.pi_args.sizes[0] = collect_pi.degree[i] * self.observation_dim
            if self.discrete:
                actors.append(CategoricalActor(**self.pi_args._toDict()))
            else:
                actors.append(GaussianActor(action_dim=self.action_dim, **self.pi_args._toDict()))
        return collect_pi, actors

    def _init_vs(self):
        collect_v = MultiCollect(torch.matrix_power(self.adj, self.radius_v), device=self.device)
        vs = nn.ModuleList()
        for i in range(self.n_agent):
            self.v_args.sizes[0] = collect_v.degree[i] * self.observation_dim
            v_fn = self.v_args.network
            vs.append(v_fn(**self.v_args._toDict()))
        return collect_v, vs

    def updateAgent(self, trajs, clip=None):
        time_t = time.time()
        if clip is None:
            clip = self.clip
        n_minibatch = self.n_minibatch

        names = Trajectory.names()
        traj_all = {name:[] for name in names}
        max_traj_length = max([i.length for i in trajs])
        for traj in trajs:
            for name in names:
                tensor_shape = traj[name].shape
                full_part_shape = [max_traj_length - tensor_shape[0]] + list(tensor_shape[1:])
                if name == 'd':
                    traj_all[name].append(torch.cat([traj[name], torch.ones(full_part_shape, dtype=torch.bool)], dim = 0))
                else:
                    traj_all[name].append(torch.cat([traj[name], torch.zeros(full_part_shape, dtype=traj[name].dtype)], dim = 0))
        # should be 4-dim [batch * step * n_agent * dim]
        traj = {name:torch.stack(value, dim=0) for name, value in traj_all.items()}


        s, a, r, s1, d, logp = traj['s'], traj['a'], traj['r'], traj['s1'], traj['d'], traj['logp']
        s, a, r, s1, d, logp = [item.to(self.device) for item in [s, a, r, s1, d, logp]]
        # all in shape [batch_size, T, n_agent, dim]
        value_old, returns, advantages, reduced_advantages = self._process_traj(**traj)

        advantages_old = reduced_advantages if self.use_reduced_v else advantages #  set use_reduced_v as False

        b, T, n, d_s = s.size()
        d_a = a.size()[-1]
        s = s.view(-1, n, d_s)
        a = a.view(-1, n, d_a)
        logp = logp.view(-1, n, 1)
        advantages_old = advantages_old.view(-1, n, 1)
        returns = returns.view(-1, n, 1)
        value_old = value_old.view(-1, n, 1)
        # s, a, logp, adv, ret, v are now all in shape [-1, n_agent, dim]

        batch_total = logp.size()[0]
        batch_size = int(batch_total/n_minibatch)

        # critic update
        for i_v in range(self.n_update_v):
            batch_returns = returns
            batch_state = s
            if n_minibatch > 1:
                idxs = np.random.randint(0, len(batch_total), size=batch_size)
                [batch_returns, batch_state] = [item[idxs] for item in [batch_returns, batch_state]]
            batch_v_new = self._evalV(batch_state)
            loss_v = ((batch_v_new - batch_returns) ** 2).mean()
            self.optimizer_v.zero_grad()
            loss_v.backward()
            self.optimizer_v.step()
            var_v = ((batch_returns - batch_returns.mean()) ** 2).mean()
            rel_v_loss = loss_v / (var_v + 1e-8)
            self.logger.log(v_loss=loss_v, v_update=None, v_var=var_v, rel_v_loss=rel_v_loss)
            if rel_v_loss < self.v_thres:
                break
        self.logger.log(v_update_step=i_v)


        # use the updated value
        _, _, advantages, _ = self._process_traj(**traj)
        advantages_old = reduced_advantages if self.use_reduced_v else advantages  # set use_reduced_v as False
        advantages_old = advantages_old.view(-1, n, 1)


        # actor update
        i_pi = 0
        for i_pi in range(self.n_update_pi):
            batch_state, batch_action, batch_logp, batch_advantages_old = [s, a, logp, advantages_old]
            if n_minibatch > 1:
                idxs = np.random.choice(range(batch_total), size=batch_size, replace=False)
                [batch_state, batch_action, batch_logp, batch_advantages_old] = [item[idxs] for item in [batch_state, batch_action, batch_logp, batch_advantages_old]]
            batch_logp_new = self.get_logp(batch_state, batch_action)

            # - A * logp - entropy_loss
            loss_pi =  torch.mean(- batch_advantages_old * batch_logp_new)
            loss_entropy = - torch.mean(batch_logp_new)
            updata_entropy_coff = max(self.entropy_coeff - self.entropy_coeff_decay * self.logger.buffer['interaction'], 0)
            loss_actor = loss_pi + loss_entropy * updata_entropy_coff
            self.optimizer_pi.zero_grad()
            loss_actor.backward()
            self.optimizer_pi.step()
            logp_diff = torch.exp(batch_logp_new) * (batch_logp_new - self.get_logp(batch_state, batch_action) )
            kl = logp_diff.mean()
            self.logger.log(pi_loss=loss_pi, entropy=loss_entropy, kl_divergence = kl, entropy_coff=updata_entropy_coff ,pi_update=None)
        self.logger.log(pi_update_step=i_pi)


        self.logger.log(update=None, reward=r, value=value_old, clip=clip, returns=returns, advantages=advantages_old.abs())
        self.logger.log(agent_update_time=time.time()-time_t)
        return [r.mean().item(), loss_entropy.item()]


class IC3Net(nn.ModuleList):
    def __init__(self, logger, device, agent_args, **kwargs):
        super().__init__()
        self.logger = logger
        self.device = device
        self.n_agent = agent_args.n_agent
        self.gamma = agent_args.gamma
        self.lamda = agent_args.lamda
        self.clip = agent_args.clip
        self.target_kl = agent_args.target_kl
        self.v_coeff = agent_args.v_coeff
        self.v_thres = agent_args.v_thres
        self.entropy_coeff = agent_args.entropy_coeff
        self.entropy_coeff_decay =  agent_args.entropy_coeff_decay  # add
        self.lr = agent_args.lr
        self.lr_v = agent_args.lr_v
        self.lr_p = agent_args.lr_p
        self.n_update_v = agent_args.n_update_v
        self.n_update_pi = agent_args.n_update_pi
        self.n_minibatch = agent_args.n_minibatch
        self.use_reduced_v = agent_args.use_reduced_v
        self.use_rtg = agent_args.use_rtg
        self.use_gae_returns = agent_args.use_gae_returns
        self.all_comm = True # TODO: need to update

        self.advantage_norm = agent_args.advantage_norm

        self.observation_dim = agent_args.observation_dim
        self.action_space = agent_args.action_space
        self.discrete = isinstance(agent_args.action_space, Discrete)
        if self.discrete:
            self.action_dim = self.action_space.n
            self.action_shape = self.action_dim
        else:
            self.action_shape = self.action_space.shape
            self.action_dim = 1
            for j in self.action_shape:
                self.action_dim *= j
            self.action_low = self.action_space.low
            self.action_high = self.action_space.high
            self.squeeze = agent_args.squeeze
        # if adj diag is not one, we should add a eye matrix
        agent_args.adj = (torch.as_tensor(agent_args.adj, device=self.device, dtype=torch.float) > 0) | torch.eye(
            self.n_agent, device=device).bool()
        self.adj = torch.as_tensor(agent_args.adj, device=self.device, dtype=torch.float)
        self.radius_v = agent_args.radius_v
        self.radius_pi = agent_args.radius_pi
        self.pi_args = agent_args.pi_args
        self.v_args = agent_args.v_args
        #self.collect_pi, self.actors = self._init_actors()
        #self.collect_v, self.vs = self._init_vs()
        self.hidden_dim = agent_args.v_args.hidden_dim


        self.initNetwork(agent_args)
        self.optimizer = Adam(list(self.obs_encoder.parameters())+list(self.comm_gate_head.parameters())+ list(self.message_models.parameters())+list(self.main_models.parameters())+list(self.value_heads.parameters())+list(self.actors.parameters()), lr=self.lr)
        #self.optimizer_pi = Adam(self.actors.parameters(), lr=self.lr)

    def initNetwork(self, agent_args):
        # [one for all]observation encoding layer
        self.obs_encoder = nn.Linear(self.observation_dim, self.hidden_dim)

        # [1 v 1] communication gated action layer, first dim for comm, second dim for not
        self.comm_gate_head = nn.ModuleList([nn.Linear(self.hidden_dim, 2) for i in range(self.n_agent)])

        # [1 v 1] message fussion layer
        self.message_models = nn.ModuleList([nn.Linear(self.hidden_dim, self.hidden_dim) for i in range(self.n_agent)])

        # [1 v 1] main layer
        self.main_models = nn.ModuleList([nn.Linear(self.hidden_dim, self.hidden_dim) for i in range(self.n_agent)])

        # value head
        self.value_heads = nn.ModuleList([nn.Linear(self.hidden_dim, 1) for i in range(self.n_agent)])

        # action head
        if self.discrete:
            self.pi_args.sizes[0] = self.hidden_dim
            self.actors = nn.ModuleList([CategoricalActor(**self.pi_args._toDict()).to(self.device) for i in range(self.n_agent)])
        else:
            self.pi_args.sizes[0] = self.hidden_dim
            self.actors = nn.ModuleList([GaussianActor(action_dim=self.action_dim, **self.pi_args._toDict()).to(self.device) for i in range(self.n_agent)])

        # activation function
        self.activation_function = torch.nn.ReLU(inplace=True) #TODO: add to config file


    def updateAgent(self, trajs, clip=None):
        time_t = time.time()
        if clip is None:
            clip = self.clip
        n_minibatch = self.n_minibatch

        names = Trajectory.names()
        traj_all = {name: [] for name in names}
        max_traj_length = max([i.length for i in trajs])
        for traj in trajs:
            for name in names:
                tensor_shape = traj[name].shape
                full_part_shape = [max_traj_length - tensor_shape[0]] + list(tensor_shape[1:])
                if name == 'd':
                    traj_all[name].append(torch.cat([traj[name], torch.ones(full_part_shape, dtype=torch.bool)], dim=0))
                else:
                    traj_all[name].append(
                        torch.cat([traj[name], torch.zeros(full_part_shape, dtype=traj[name].dtype)], dim=0))
        # should be 4-dim [batch * step * n_agent * dim]
        traj = {name: torch.stack(value, dim=0) for name, value in traj_all.items()}

        s, a, r, s1, d, logp = traj['s'], traj['a'], traj['r'], traj['s1'], traj['d'], traj['logp']
        s, a, r, s1, d, logp = [item.to(self.device) for item in [s, a, r, s1, d, logp]]
        # all in shape [batch_size, T, n_agent, dim]
        value_old, returns, advantages, reduced_advantages = self._process_traj(**traj)

        advantages_old = reduced_advantages if self.use_reduced_v else advantages  # set use_reduced_v as False

        b, T, n, d_s = s.size()
        d_a = a.size()[-1]
        s = s.view(-1, n, d_s)
        a = a.view(-1, n, d_a)
        logp = logp.view(-1, n, 1)
        advantages_old = advantages_old.view(-1, n, 1)
        returns = returns.view(-1, n, 1)
        value_old = value_old.view(-1, n, 1)
        # s, a, logp, adv, ret, v are now all in shape [-1, n_agent, dim]

        batch_total = logp.size()[0]
        batch_size = int(batch_total / n_minibatch)

        # critic loss
        batch_returns = returns
        batch_state = s
        if n_minibatch > 1:
            idxs = np.random.randint(0, len(batch_total), size=batch_size)
            [batch_returns, batch_state] = [item[idxs] for item in [batch_returns, batch_state]]
        batch_v_new = self._evalV(batch_state)
        loss_v = ((batch_v_new - batch_returns) ** 2).mean()
        var_v = ((batch_returns - batch_returns.mean()) ** 2).mean()
        rel_v_loss = loss_v / (var_v + 1e-8)
        self.logger.log(v_loss=loss_v, v_update=None, v_var=var_v, rel_v_loss=rel_v_loss)
        # if rel_v_loss < self.v_thres:
        #     break
        self.logger.log(v_update_step=1)

        # actor loss
        batch_state, batch_action, batch_logp, batch_advantages_old = [s, a, logp, advantages_old]
        if n_minibatch > 1:
            idxs = np.random.choice(range(batch_total), size=batch_size, replace=False)
            [batch_state, batch_action, batch_logp, batch_advantages_old] = [item[idxs] for item in
                                                                             [batch_state, batch_action, batch_logp,
                                                                              batch_advantages_old]]
        batch_logp_new = self.get_logp(batch_state, batch_action)

        # - A * logp - entropy_loss
        loss_pi = torch.mean(- batch_advantages_old * batch_logp_new)
        loss_entropy = - torch.mean(batch_logp_new)
        updata_entropy_coff = max(self.entropy_coeff - self.entropy_coeff_decay * self.logger.buffer['interaction'],
                                  0)
        loss_actor = loss_pi + loss_entropy * updata_entropy_coff

        loss = self.lr_v * loss_v + self.lr_p * loss_actor

        self.optimizer.zero_grad()
        loss.backward()
        self.optimizer.step()

        logp_diff = torch.exp(batch_logp_new) * (batch_logp_new - self.get_logp(batch_state, batch_action))
        kl = logp_diff.mean()
        self.logger.log(pi_loss=loss_pi, entropy=loss_entropy, kl_divergence=kl, entropy_coff=updata_entropy_coff,
                        pi_update=None)
        self.logger.log(pi_update_step=1)

    def checkConverged(self, ls_info):
        #TODO: not neccessary
        return False

    def group_inference(self, model, data):
        # model is a modelList
        # data is a [batch * n_agent *dim]
        outs = []
        for i in range(self.n_agent):
            agent_data = data.select(1, i)
            outs.append(model[i](agent_data))
        outs = torch.stack(outs, dim=1)
        return outs

    def inference_hidden_state(self, s):
        # encode the state
        s_encoding = self.activation_function(self.obs_encoder(s))

        # decide which agent to communication
        # [batch_size, n_agent, 2]
        comm_gate_distirbution = self.group_inference(self.comm_gate_head, s_encoding)


        # merge the message by mean
        if self.all_comm == True:
            batch_comm_adj = self.adj.unsqueeze(0).repeat(s.shape[0], 1, 1)
        else:
            # TODO: check if need detach
            comm_gate = torch.stack(
                [torch.stack([torch.multinomial(dis, 1).detach() for dis in n_dis]) for n_dis in
                 comm_gate_distirbution])
            comm_gate_tensor = comm_gate.unsqueeze(2).repeat(1, 1, self.n_agent)
            comm_gate_tensor_T = comm_gate_tensor.permute(0, 2, 1)
            comm_gate = comm_gate_tensor * comm_gate_tensor_T
            batch_comm_adj = self.adj.unsqueeze(0).repeat(comm_gate.shape[0], 1, 1) | comm_gate

        batch_message_ls = []
        for i in range(s.shape[0]):
            # for each batch
            comm_adj = batch_comm_adj.select(0, i)
            single_s_encoding = s_encoding.select(0, i)

            batch_message_ls.append(torch.mm(comm_adj, single_s_encoding))
        batch_message = torch.stack(batch_message_ls, dim=0)

        # deal with the meassage
        deal_message = self.group_inference(self.message_models, batch_message)

        # genarate hidden state
        hidden_state = s_encoding + self.group_inference(self.main_models, s_encoding) + deal_message
        return hidden_state

    def save(self, info=None):
        self.logger.save(self, info=info)


    def act(self, s, requires_log=False):
        """
                Requires input of [batch_size, n_agent, dim] or [n_agent, dim].
                This method is gradient-free. To get the gradient-enabled probability information, use get_logp().
                Returns a distribution with the same dimensions of input.
                """
        with torch.no_grad():
            dim = s.dim()
            while s.dim() <= 2:
                s = s.unsqueeze(0)
            s = s.to(self.device)

            hidden_state = self.inference_hidden_state(s)
            hidden_state =  hidden_state.permute(1, 0, 2) # Now s[i].dim() == 2 ([batch_size, dim])

            # cal the action
            if self.discrete:
                probs = []
                for i in range(self.n_agent):
                    probs.append(self.actors[i](hidden_state[i]))
                probs = torch.stack(probs, dim=1)
                return Categorical(probs)
            else:
                means, stds = [], []
                for i in range(self.n_agent):
                    mean, std = self.actors[i](hidden_state[i])
                    means.append(mean)
                    stds.append(std)
                means = torch.stack(means, dim=1)
                stds = torch.stack(stds, dim=1)
                while means.dim() > dim:
                    means = means.squeeze(0)
                    stds = stds.squeeze(0)
                return Normal(means, stds)
                # return means

    def get_logp(self, s, a):
        s = torch.as_tensor(s, dtype=torch.float32, device=self.device)
        dim = s.dim()
        while s.dim() <= 2:
            s = s.unsqueeze(0)
            a = a.unsqueeze(0)
        while a.dim() < s.dim():
            a = a.unsqueeze(-1)

        hidden_state = self.inference_hidden_state(s)
        hidden_state = hidden_state.permute(1, 0, 2)  # Now s[i].dim() == 2 ([batch_size, dim])

        log_prob = []
        for i in range(self.n_agent):
            if self.discrete:
                probs = self.actors[i](hidden_state[i])
                log_prob.append(torch.log(torch.gather(probs, dim=-1, index=torch.select(a, dim=1, index=i).long())))
            else:
                log_prob.append(self.actors[i](hidden_state[i], a.select(dim=1, index=i)))
        log_prob = torch.stack(log_prob, dim=1)
        while log_prob.dim() < 3:
            log_prob = log_prob.unsqueeze(-1)
        return log_prob

    def _evalV(self, s):
        # TODO￡o
        # Requires input in shape [-1, n_agent, dim]
        s = s.to(self.device)

        hidden_state = self.inference_hidden_state(s)
        hidden_state = hidden_state.permute(1, 0, 2)  # Now s[i].dim() == 2 ([batch_size, dim])

        values = []
        for i in range(self.n_agent):
            values.append(self.value_heads[i](hidden_state[i]))
        return torch.stack(values, dim=1)

    def _process_traj(self, s, a, r, s1, d, logp):
        """
        Input are all in shape [batch_size, T, n_agent, dim]
        """
        with torch.no_grad():
            b, T, n, dim_s = s.shape
            s, a, r, s1, d, logp = [item.to(self.device) for item in [s, a, r, s1, d, logp]]
            value = self._evalV(s.view(-1, n, dim_s)).view(b, T, n, -1)

            returns = torch.zeros(value.size(), device=self.device)
            deltas, advantages = torch.zeros_like(returns), torch.zeros_like(returns)
            prev_value = self._evalV(s1.select(1, T - 1))
            if not self.use_rtg:
                prev_return = prev_value
            else:
                prev_return = torch.zeros_like(prev_value)
            prev_advantage = torch.zeros_like(prev_return)
            d_mask = d.float()
            for t in reversed(range(T)):
                deltas[:, t, :, :]= r.select(1, t) + self.gamma * (1-d_mask.select(1,t)) * prev_value - value.select(1, t).detach()
                advantages[:, t, :, :] = deltas.select(1, t) + self.gamma * self.lamda * (1-d_mask.select(1,t)) * prev_advantage
                if self.use_gae_returns:
                    returns[:, t, :, :] = value.select(1, t).detach() + advantages.select(1, t)
                else:
                    returns[:, t, :, :] = r.select(1, t) + self.gamma * (1-d_mask.select(1, t)) * prev_return

                prev_return = returns.select(1, t)
                prev_value = value.select(1, t)
                prev_advantage = advantages.select(1, t)

        return value, returns, advantages, None

class DPPOAgent(nn.ModuleList):
    """
    Everything in and out is torch Tensor.
    """
    def __init__(self, logger, device, agent_args, **kwargs):
        super().__init__()
        self.logger = logger
        self.device = device
        self.n_agent = agent_args.n_agent
        self.gamma = agent_args.gamma
        self.lamda = agent_args.lamda
        self.clip = agent_args.clip
        self.target_kl = agent_args.target_kl
        self.v_coeff = agent_args.v_coeff
        self.v_thres = agent_args.v_thres
        self.entropy_coeff = agent_args.entropy_coeff
        self.lr = agent_args.lr
        self.lr_v = agent_args.lr_v
        self.n_update_v = agent_args.n_update_v
        self.n_update_pi = agent_args.n_update_pi
        self.n_minibatch = agent_args.n_minibatch
        self.use_reduced_v = agent_args.use_reduced_v
        self.use_rtg = agent_args.use_rtg
        self.use_gae_returns = agent_args.use_gae_returns

        self.advantage_norm = agent_args.advantage_norm

        self.observation_dim = agent_args.observation_dim
        self.action_space = agent_args.action_space
        self.discrete = isinstance(agent_args.action_space, Discrete)
        
        if self.discrete:
            self.action_dim = self.action_space.n
            self.action_shape = self.action_dim
            
            
        # else:
        #     self.action_shape = self.action_space.shape
        #     # print(self.action_shape)
        #     # self.action_dim = self.action_space.shape[0]
        #     self.action_dim = 2
        #     for j in self.action_shape:
        #         self.action_dim *= j
        #     self.action_low = self.action_space.low.item()
        #     self.action_high = self.action_space.high.item()
        #     self.squeeze = agent_args.squeeze
        
            
        else:
            self.action_shape = self.action_space.shape
            self.action_dim = self.action_space.shape[0]
            # for j in self.action_shape:
            #     print("j=",j)
            #     self.action_dim *= j
            self.action_low = self.action_space.low
            self.action_high = self.action_space.high
            self.squeeze = agent_args.squeeze
            
        
        self.adj = torch.as_tensor(agent_args.adj, device=self.device, dtype=torch.float)
        self.radius_v = agent_args.radius_v
        self.radius_pi = agent_args.radius_pi
        self.pi_args = agent_args.pi_args
        self.v_args = agent_args.v_args
        self.collect_pi, self.actors = self._init_actors()
        
        #load_net-----------------------------------------------------------

        # self.actors.load_state_dict(torch.load(test_actors_model, map_location={'cuda:5':'cuda:0'}))      
        # self.actors.load_state_dict(torch.load(test_actors_model, map_location={'cuda:1':'cuda:0'}))    
        # self.actors.load_state_dict(torch.load(test_actors_model))  
        #-------------------------------------------------------------------
        self.actors.load_state_dict(torch.load(test_actors_model, map_location={'cuda:0':'cuda:0','cuda:1':'cuda:0','cuda:2':'cuda:0','cuda:3':'cuda:0','cuda:4':'cuda:0','cuda:5':'cuda:0',}))    

        
        self.collect_v, self.vs = self._init_vs()

        self.optimizer_v = Adam(self.vs.parameters(), lr=self.lr_v)
        self.optimizer_pi = Adam(self.actors.parameters(), lr=self.lr)

    def act(self, s, requires_log=False):
        """
        Requires input of [batch_size, n_agent, dim] or [n_agent, dim].
        This method is gradient-free. To get the gradient-enabled probability information, use get_logp().
        Returns a distribution with the same dimensions of input.
        """
        if_test=False
        with torch.no_grad():
            dim = s.dim()
            while s.dim() <= 2:
                s = s.unsqueeze(0)
            s = s.to(self.device)
            s = self.collect_pi.gather(s) # all state into [ self +  ]
            # Now s[i].dim() == 2 ([batch_size, dim])


            if self.discrete:
                if if_test: 
                    probs = []
                    for i in range(self.n_agent):
                        index = torch.argmax(self.actors[i](s[i])[0])
                        action_p = torch.tensor([[0]*self.action_dim])
                        action_p[0][index] = 1
                        probs.append(action_p)
                    probs = torch.stack(probs, dim=1)
                    #print("probs=",probs)
                    while probs.dim() > dim:
                        probs = probs.squeeze(0)
                    return Categorical(probs)
                  
                else:
                    probs = []
                    for i in range(self.n_agent):
                        probs.append(self.actors[i](s[i]))
                    probs = torch.stack(probs, dim=1)
                    while probs.dim() > dim:
                        probs = probs.squeeze(0)
                    return Categorical(probs)



            else:
                means, stds = [], []
                for i in range(self.n_agent):
                    mean, std = self.actors[i](s[i])
                    means.append(mean)
                    

                    stds.append(std)
                    # stds.append(std.exp())
                    
                    
                    
                means = torch.stack(means, dim=1)
                stds = torch.stack(stds, dim=1)
                while means.dim() > dim:
                    means = means.squeeze(0)
                    stds = stds.squeeze(0)
                return Normal(means, stds)
    
    def get_logp(self, s, a):
        """
        Requires input of [batch_size, n_agent, dim] or [n_agent, dim].
        Returns a tensor whose dim() == 3.
        """
        s = torch.as_tensor(s, dtype=torch.float32, device=self.device)
        dim = s.dim()

        while s.dim() <= 2:
            s = s.unsqueeze(0)
            a = a.unsqueeze(0)
        while a.dim() < s.dim():
            a = a.unsqueeze(-1)
            
        s = self.collect_pi.gather(s)
        # Now s[i].dim() == 2, a.dim() == 3
        log_prob = []
        for i in range(self.n_agent):
            if self.discrete:
                probs = self.actors[i](s[i])
                log_prob.append(torch.log(torch.gather(probs, dim=-1, index=torch.select(a, dim=1, index=i).long())))
            else:
                log_prob.append(self.actors[i](s[i], a.select(dim=1, index=i)))
        log_prob = torch.stack(log_prob, dim=1)
        while log_prob.dim() < 3:
            log_prob = log_prob.unsqueeze(-1)
        return log_prob

    def updateAgent(self, trajs, clip=None):
        time_t = time.time()
        if clip is None:
            clip = self.clip
        n_minibatch = self.n_minibatch

        names = Trajectory.names()
        traj_all = {name:[] for name in names}
        max_traj_length = max([i.length for i in trajs])
        for traj in trajs:
            for name in names:
                tensor_shape = traj[name].shape
                full_part_shape = [max_traj_length - tensor_shape[0]] + list(tensor_shape[1:])
                if name == 'd':
                    traj_all[name].append(torch.cat([traj[name], torch.ones(full_part_shape, dtype=torch.bool)], dim=0))
                else:
                    traj_all[name].append(torch.cat([traj[name], torch.zeros(full_part_shape, dtype=traj[name].dtype)], dim=0))
        traj = {name:torch.stack(value, dim=0) for name, value in traj_all.items()}

        for i_update in range(self.n_update_pi):
            s, a, r, s1, d, logp = traj['s'], traj['a'], traj['r'], traj['s1'], traj['d'], traj['logp']
            s, a, r, s1, d, logp = [item.to(self.device) for item in [s, a, r, s1, d, logp]]
            # all in shape [batch_size, T, n_agent, dim]
            value_old, returns, advantages, reduced_advantages = self._process_traj(**traj)

            advantages_old = reduced_advantages if self.use_reduced_v else advantages

            b, T, n, d_s = s.size()

            d_a = a.size()[-1]
            d_p = logp.size()[-1]
            
            s = s.view(-1, n, d_s)
            a = a.view(-1, n, d_a)
            logp = logp.view(-1, n, d_p)    # for DMPO            
            #-------------------------------------------------------------------------
            # logp = np.squeeze(logp, axis=0)   #for uav_env

            #-------------------------------------------------------------------------     
            
            
            # logp = logp.view(-1, n, 1)    #pre_code

            advantages_old = advantages_old.view(-1, n, 1)
            returns = returns.view(-1, n, 1)
            value_old = value_old.view(-1, n, 1)
            # s, a, logp, adv, ret, v are now all in shape [-1, n_agent, dim]

            batch_total = logp.size()[0]
            batch_size = int(batch_total/n_minibatch)
            kl_all = []
            i_pi = 0
            for i_pi in range(1):  
                batch_state, batch_action, batch_logp, batch_advantages_old = [s, a, logp, advantages_old]
                if n_minibatch > 1:
                    idxs = np.random.choice(range(batch_total), size=batch_size, replace=False)
                    [batch_state, batch_action, batch_logp, batch_advantages_old] = [item[idxs] for item in [batch_state, batch_action, batch_logp, batch_advantages_old]]
                batch_logp_new = self.get_logp(batch_state, batch_action)
                

                
                logp_diff = batch_logp_new - batch_logp
                kl = logp_diff.mean()
                ratio = torch.exp(batch_logp_new - batch_logp)
                surr1 = ratio * batch_advantages_old
                surr2 = ratio.clamp(1 - clip, 1 + clip) * batch_advantages_old
                loss_surr = torch.min(surr1, surr2).mean()
                loss_entropy = - torch.mean(batch_logp_new)
                loss_pi = - loss_surr - self.entropy_coeff * loss_entropy
                self.optimizer_pi.zero_grad()
                loss_pi.backward()
                self.optimizer_pi.step()
                self.logger.log(surr_loss = loss_surr, entropy = loss_entropy, kl_divergence = kl, pi_update=None)
                kl_all.append(kl.abs().item())
                if self.target_kl is not None and kl.abs() > 1.5 * self.target_kl:
                    break
            self.logger.log(pi_update_step=i_update)

            for i_v in range(1):
                batch_returns = returns
                batch_state = s
                if n_minibatch > 1:
                    idxs = np.random.randint(0, len(batch_total), size=batch_size)
                    [batch_returns, batch_state] = [item[idxs] for item in [batch_returns, batch_state]]
                batch_v_new = self._evalV(batch_state)
                loss_v = ((batch_v_new - batch_returns) ** 2).mean()
                self.optimizer_v.zero_grad()
                loss_v.backward()
                self.optimizer_v.step()
                var_v = ((batch_returns - batch_returns.mean()) ** 2).mean()
                rel_v_loss = loss_v / (var_v + 1e-8)
                self.logger.log(v_loss=loss_v, v_update=None, v_var=var_v, rel_v_loss=rel_v_loss)
                if rel_v_loss < self.v_thres:
                    break
            self.logger.log(v_update_step=i_update)
            self.logger.log(update=None, reward=r, value=value_old, clip=clip, returns=returns, advantages=advantages_old.abs())
        self.logger.log(agent_update_time=time.time()-time_t)
        return [r.mean().item(), loss_entropy.item(), max(kl_all)]
    
    def checkConverged(self, ls_info):
        return False

    def save(self, info=None):
        self.logger.save(self, info=info)

    def load(self, state_dict):
        self.load_state_dict(state_dict[self.logger.prefix])

#_____________________________________________________________________________________________________--

    def save_nets(self, dir_name,episode):
        if not os.path.exists(dir_name + '/Models'):
            os.mkdir(dir_name + '/Models')
        # torch.save(self.critic.state_dict(), dir_name + '/Models/' +str(episode)+ 'best_critic.pt')
        torch.save(self.actors.state_dict(), dir_name + '/Models/' + str(episode)+ 'best_actor.pt')
        # torch.save(self.actors.state_dict(), dir_name + '/' +str(episode)+ 'best_actor.pt')
        print('RL saved successfully')

    def load_nets(self, dir_name,episode):
        
        self.actors.load_state_dict(torch.load(dir_name + '/Models/' + str(episode)+ 'best_actor.pt'))


        # print('RL load successfully')
#_____________________________________________________________________________________________________--

    def _evalV(self, s):
        # Requires input in shape [-1, n_agent, dim]
        s = s.to(self.device)
        s = self.collect_v.gather(s)
        values = []
        for i in range(self.n_agent):
            values.append(self.vs[i](s[i]))
        return torch.stack(values, dim=1)

    def _init_actors(self):
        collect_pi = MultiCollect(torch.matrix_power(self.adj, self.radius_pi), device=self.device)
        actors = nn.ModuleList()
        for i in range(self.n_agent):
            self.pi_args.sizes[0] = collect_pi.degree[i] * self.observation_dim
            if self.discrete:
                actors.append(CategoricalActor(**self.pi_args._toDict()).to(self.device))
            else:
                actors.append(GaussianActor(action_dim=self.action_dim, **self.pi_args._toDict()).to(self.device))
        
        return collect_pi, actors
    
    def _init_vs(self):
        collect_v = MultiCollect(torch.matrix_power(self.adj, self.radius_v), device=self.device)
        vs = nn.ModuleList()
        for i in range(self.n_agent):
            self.v_args.sizes[0] = collect_v.degree[i] * self.observation_dim
            v_fn = self.v_args.network
            vs.append(v_fn(**self.v_args._toDict()).to(self.device))
        return collect_v, vs
    
    def _process_traj(self, s, a, r, s1, d, logp):
        """
        Input are all in shape [batch_size, T, n_agent, dim]
        """
        b, T, n, dim_s = s.shape
        s, a, r, s1, d, logp = [item.to(self.device) for item in [s, a, r, s1, d, logp]]
        value = self._evalV(s.view(-1, n, dim_s)).view(b, T, n, -1)

        returns = torch.zeros(value.size(), device=self.device)
        deltas, advantages = torch.zeros_like(returns), torch.zeros_like(returns)
        prev_value = self._evalV(s1.select(1, T - 1))
        if not self.use_rtg:
            prev_return = prev_value
        else:
            prev_return = torch.zeros_like(prev_value)
        prev_advantage = torch.zeros_like(prev_return)
        d_mask = d.float()
        for t in reversed(range(T)):
            deltas[:, t, :, :]= r.select(1, t) + self.gamma * (1-d_mask.select(1,t)) * prev_value - value.select(1, t).detach()
            advantages[:, t, :, :] = deltas.select(1, t) + self.gamma * self.lamda * (1-d_mask.select(1,t)) * prev_advantage
            if self.use_gae_returns:
                returns[:, t, :, :] = value.select(1, t).detach() + advantages.select(1, t)
            else:
                returns[:, t, :, :] = r.select(1, t) + self.gamma * (1-d_mask.select(1, t)) * prev_return

            prev_return = returns.select(1, t)
            prev_value = value.select(1, t)
            prev_advantage = advantages.select(1, t)
        reduced_advantages = self.collect_v.reduce_sum(advantages.view(-1, n, 1)).view(advantages.size())
        if self.advantage_norm and reduced_advantages.size()[1] > 1:
            reduced_advantages = (reduced_advantages - reduced_advantages.mean(dim=1, keepdim=True)) / (reduced_advantages.std(dim=1, keepdim=True) + 1e-5)
            advantages = (advantages - advantages.mean(dim=1, keepdim=True)) / (advantages.std(dim=1, keepdim=True) + 1e-5)
        return value.detach(), returns, advantages.detach(), reduced_advantages.detach()

class ModelBasedAgent(nn.ModuleList):
    def __init__(self, logger, device, agent_args, **kwargs):
        super().__init__(logger, device, agent_args, **kwargs)
        self.logger = logger
        self.device = device
        self.lr_p = agent_args.lr_p
        self.p_args = agent_args.p_args
        self.ps = GraphConvolutionalModel(self.logger, self.adj, self.observation_dim, self.action_dim, self.n_agent, self.p_args).to(self.device)
        self.optimizer_p = Adam(self.ps.parameters(), lr=self.lr)

    def updateModel(self, trajs, length=1):
        """
        Input dim: 
        s: [[T, n_agent, state_dim]]
        a: [[T, n_agent, action_dim]]
        """
        time_t = time.time()
        loss_total = 0.
        ss, actions, rs, s1s, ds = [], [], [], [], []
        for traj in trajs:
            s, a, r, s1, d = traj["s"], traj["a"], traj["r"], traj["s1"], traj["d"]
            s, a, r, s1, d = [torch.as_tensor(item, device=self.device) for item in [s, a, r, s1, d]]
            ss.append(s)
            actions.append(a)
            rs.append(r)
            s1s.append(s1)
            ds.append(d)
        ss, actions, rs, s1s, ds = [torch.stack(item, dim=0) for item in [ss, actions, rs, s1s, ds]]
        loss, rel_state_error = self.ps.train(ss, actions, rs, s1s, ds, length) # [n_traj, T, n_agent, dim]
        self.optimizer_p.zero_grad()
        loss.sum().backward()
        # torch.nn.utils.clip_grad_norm_(parameters=self.ps.parameters(), max_norm=5, norm_type=2)
        self.optimizer_p.step()
        self.logger.log(p_loss_total=loss.sum(), p_update=None)
        self.logger.log(model_update_time=time.time()-time_t)
        return rel_state_error.item()
    
    def validateModel(self, trajs, length=1):
        with torch.no_grad():
            ss, actions, rs, s1s, ds = [], [], [], [], []
            for traj in trajs:
                s, a, r, s1, d = traj["s"], traj["a"], traj["r"], traj["s1"], traj["d"]
                s, a, r, s1, d = [torch.as_tensor(item, device=self.device) for item in [s, a, r, s1, d]]
                ss.append(s)
                actions.append(a)
                rs.append(r)
                s1s.append(s1)
                ds.append(d)
            ss, actions, rs, s1s, ds = [torch.stack(item, dim=0) for item in [ss, actions, rs, s1s, ds]]
            _, rel_state_error = self.ps.train(ss, actions, rs, s1s, ds, length) # [n_traj, T, n_agent, dim]
            return rel_state_error.item()
    
    def model_step(self, s, a):
        """
        Input dim: 
        s: [batch_size, n_agent, state_dim]
        a: [batch_size, n_agent] (discrete) or [batch_size, n_agent, action_dim] (continuous)

        Return dim == 3.
        """
        with torch.no_grad():
            while s.dim() <= 2:
                s = s.unsqueeze(0)
                a = a.unsqueeze(0)
            while a.dim() <= 2:
                a = a.unsqueeze(-1)
            s = s.to(self.device)
            a = a.to(self.device)
            rs, s1s, ds = self.ps.predict(s, a)
            return rs.detach(), s1s.detach(), ds.detach(), s.detach()
    
    def load_model(self, pretrained_model):
        dic = torch.load(pretrained_model)
        self.load_state_dict(dic[''])



class HiddenAgent(ModelBasedAgent):
    def __init__(self, logger, device, agent_args, **kwargs):
        super().__init__(logger, device, agent_args, **kwargs)
        self.hidden_state_dim = agent_args.hidden_state_dim
        self.embedding_sizes = agent_args.embedding_sizes
        self.embedding_layers = self._init_embedding_layers()
        self.optimizer_p.add_param_group({'params': self.embedding_layers.parameters()})
    
    def act(self, s, requires_log=False):
        s = s.detach()
        if s.size()[-1] != self.hidden_state_dim:
            s = self._state_embedding(s).detach()
        return super().act(s, requires_log)
    
    def get_logp(self, s, a):
        s = s.detach()
        if s.size()[-1] != self.hidden_state_dim:
            s = self._state_embedding(s).detach()
        return super().get_logp(s, a)
    
    def updateModel(self, s, a, r, s1, d):
        if s.size()[-1] != self.hidden_state_dim:
            s = self._state_embedding(s)
        if s1.size()[-1] != self.hidden_state_dim:
            s1 = self._state_embedding(s1)
        return super().updateModel(s, a, r, s1, d)
    
    def model_step(self, s, a):
        if s.size()[-1] != self.hidden_state_dim:
            s = self._state_embedding(s)
        return super().model_step(s, a)

    def _init_embedding_layers(self):
        embedding_layers = nn.ModuleList()
        for _ in range(self.n_agent):
            embedding_layers.append(MLP(self.embedding_sizes, activation=nn.ReLU))
        return embedding_layers.to(self.device)
    
    def _state_embedding(self, s):
        embeddings = []
        for i in range(self.n_agent):
            embeddings.append(self.embedding_layers[i](s.select(dim=-2, index=i).to(self.device)))
        embeddings = torch.stack(embeddings, dim=-2)
        return embeddings

class MB_DPPOAgent(ModelBasedAgent, DPPOAgent):
    def __init__(self, logger, device, agent_args, **kwargs):
        super().__init__(logger, device, agent_args, **kwargs)
    
    def checkConverged(self, ls_info):
        rs = [info[0] for info in ls_info]
        r_converged = len(rs) > 8 and np.mean(rs[-3:]) < np.mean(rs[:-5])
        entropies = [info[1] for info in ls_info]
        entropy_converged = len(entropies) > 8 and np.abs(np.mean(entropies[-3:]) / np.mean(entropies[:-5]) - 1) < 1e-2
        kls = [info[2] for info in ls_info]
        kl_exceeded = False
        if self.target_kl is not None:
            kls = [kl > 1.5 * self.target_kl for kl in kls]
            kl_exceeded = any(kls)
        return kl_exceeded or r_converged and entropy_converged

class MB_DPPOAgent_Hidden(HiddenAgent, MB_DPPOAgent):
    def __init__(self, logger, device, agent_args, **kwargs):
        super().__init__(logger, device, agent_args, **kwargs)
    
    def checkConverged(self, ls_info):
        rs = [info[0] for info in ls_info]
        r_converged = len(rs) > 8 and np.mean(rs[-3:]) < np.mean(rs[:-5])
        entropies = [info[1] for info in ls_info]
        entropy_converged = len(entropies) > 8 and np.abs(np.mean(entropies[-3:]) / np.mean(entropies[:-5]) - 1) < 1e-2
        kls = [info[2] for info in ls_info]
        kl_exceeded = False
        if self.target_kl is not None:
            kls = [kl > 1.5 * self.target_kl for kl in kls]
            kl_exceeded = any(kls)
        return kl_exceeded or r_converged and entropy_converged

"""
class DA2CAgent(DPPOAgent):
    def __init__(self, logger, device, agent_args, **kwargs):
        super().__init__(logger, device, agent_args, **kwargs)
        self.use_rtg = False

    def updateAgent(self, traj, clip=None):
        time_t = time.time()
        if clip is None:
            clip = self.clip
        n_minibatch = self.n_minibatch
        s, a, r, s1, d, logp = traj['s'], traj['a'], traj['r'], traj['s1'], traj['d'], traj['logp']
        s, a, r, s1, d, logp = [item.to(self.device) for item in [s, a, r, s1, d, logp]]
        value_old, returns, advantages, reduced_advantages = self._process_traj(traj)

        advantages_old = reduced_advantages if self.use_reduced_v else advantages

        b, T, n, d_s = s.size()
        d_a = a.size()[-1]
        s = s.view(-1, n, d_s)
        a = a.view(-1, n, d_a)
        logp = logp.view(-1, n, 1)
        advantages_old = advantages_old.view(-1, n, 1)
        returns = returns.view(-1, n, 1)
        value_old = value_old.view(-1, n, 1)

        batch_total = logp.size()[0]
        batch_size = int(batch_total/n_minibatch)
        kl = 0
        i_pi = 0
        for i_pi in range(self.n_update_pi):
            batch_state, batch_action, batch_logp, batch_advantages_old = [s, a, logp, advantages_old]
            if n_minibatch > 1:
                idxs = np.random.choice(range(batch_total), size=batch_size, replace=False)
                [batch_state, batch_action, batch_logp, batch_advantages_old] = [item[idxs] for item in [batch_state, batch_action, batch_logp, batch_advantages_old]]
            batch_logp_new = self.get_logp(batch_state, batch_action)
            logp_diff = batch_logp_new - batch_logp
            kl = logp_diff.mean()
            loss_surr = batch_logp_new * batch_advantages_old
            loss_entropy = - torch.mean(torch.exp(batch_logp_new) * batch_logp_new)
            loss_pi = - loss_surr - self.entropy_coeff * loss_entropy
            self.optimizer_pi.zero_grad()
            loss_pi.backward()
            self.optimizer_pi.step()
            self.logger.log(surr_loss = loss_surr, entropy = loss_entropy, kl_divergence = kl, pi_update=None)
            if self.target_kl is not None and kl.abs() > 1.5 * self.target_kl:
                break
        self.logger.log(pi_update_step=i_pi)

        for _ in range(self.n_update_v):
            batch_returns = returns
            batch_state = s
            if n_minibatch > 1:
                idxs = np.random.randint(0, len(batch_total), size=batch_size)
                [batch_returns, batch_state] = [item[idxs] for item in [batch_returns, batch_state]]
            batch_v_new = self._evalV(batch_state)
            loss_v = ((batch_v_new - batch_returns) ** 2).mean()
            self.optimizer_v.zero_grad()
            loss_v.backward()
            self.optimizer_v.step()
            self.logger.log(v_loss=loss_v, v_update=None)
        self.logger.log(update=None, reward=r, value=value_old, clip=clip, returns=returns, advantages=advantages_old.abs())
        self.logger.log(agent_update_time=time.time()-time_t)
"""