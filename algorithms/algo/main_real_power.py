import time
import os
from numpy.core.numeric import indices
from torch.distributions.normal import Normal
from algorithms.utils import collect, mem_report
from algorithms.models import GaussianActor, GraphConvolutionalModel, MLP, CategoricalActor
from tqdm.std import trange
#from algorithms.algorithm import ReplayBuffer
#from ray.state import actors
from gym.spaces.box import Box
from gym.spaces.discrete import Discrete
import torch
import torch.nn as nn
from torch.distributions.categorical import Categorical
from torch.optim import Adam
import numpy as np
import pickle
from copy import deepcopy as dp
from algorithms.models import CategoricalActor, EnsembledModel, SquashedGaussianActor, ParameterizedModel_MBPPO
import random
import multiprocessing as mp
# import torch.multiprocessing as mp
from torch import distributed as dist
import argparse
from algorithms.algo.buffer import MultiCollect,Trajectory,TrajectoryBuffer,ModelBuffer
from algorithms.algo.normalization_utils import ZFilter, RunningStat


def translate_action(action_bias, action_scale, action):
    action = torch.as_tensor(action, dtype=torch.float)
    actions = action.detach().squeeze()
    # clip and scale action to correct range for safety
    cp_actions = torch.clamp(actions, min=-1.0, max=1.0)
    #cp_actions = torch.tanh(actions)
    low = action_bias - action_scale
    high = action_bias + action_scale
    cp_actions = 0.5 * (cp_actions + 1.0) * (high - low) + low
    cp_actions = cp_actions.cpu().numpy()
    return cp_actions

def translate_action_2(action_bias, action_scale, action):
    actions = action.detach().squeeze()
    # clip and scale action to correct range for safety
    cp_actions = torch.clamp(actions, min=-1.0, max=1.0)
    #cp_actions = torch.tanh(actions)
    low = action_bias - action_scale
    high = action_bias + action_scale
    cp_actions = 0.5 * (cp_actions + 1.0) * (high - low) + low
    #cp_actions = cp_actions.cpu().numpy()
    return cp_actions


def transfer_action_real_power(a,result):
    b = np.array([])
    
    for i in range(result.shape[0]):
        row = a[i, :]
        mask = result[i, :]
        non_zeros = row[np.nonzero(mask)]  # 当前行 mask 中值为非零的元素
        res = a[i, len(non_zeros):]
        if len(res)!=0:
            b = np.concatenate([b,np.mean(res) + non_zeros])
        else:
            b = np.concatenate([b,non_zeros])
    return b

class OnPolicyRunner:
    def __init__(self, logger, run_args, alg_args, agent, env_learn, env_test, env_args,**kwargs):
        self.logger = logger
        self.name = run_args.name
        if not run_args.init_checkpoint is None:
            agent.load(run_args.init_checkpoint)
            logger.log(interaction=run_args.start_step)  
        self.start_step = run_args.start_step 
        self.env_name = env_args.env
        self.algo_name = env_args.algo
        

        # algorithm arguments
        self.n_iter = alg_args.n_iter
        self.n_inner_iter = alg_args.n_inner_iter
        self.n_warmup = alg_args.n_warmup
        self.n_model_update = alg_args.n_model_update
        self.n_model_update_warmup = alg_args.n_model_update_warmup
        self.n_test = alg_args.n_test
        self.test_interval = alg_args.test_interval
        self.rollout_length = alg_args.rollout_length
        self.test_length = alg_args.test_length
        self.max_episode_len = alg_args.max_episode_len
        self.clip_scheme = None if (not hasattr(alg_args, "clip_scheme")) else alg_args.clip_scheme
        
        # agent initialization
        self.agent = agent
        self.device = self.agent.device if hasattr(self.agent, "device") else "cpu"

        # environment initialization
        self.env_learn = env_learn
        self.env_test = env_test
        if self.env_name == 'PowerGrid' and self.env_learn.n_agent==40:
            self.running_state = ZFilter((self.env_learn.n_agent,self.env_learn.n_s), clip=5.0)

        

        # buffer initialization
        self.discrete = agent.discrete
        action_dtype = torch.long if self.discrete else torch.float
        self.model_based = alg_args.model_based
        self.model_batch_size = alg_args.model_batch_size
        if self.model_based:
            self.n_traj = alg_args.n_traj
            self.model_traj_length = alg_args.model_traj_length
            self.model_error_thres = alg_args.model_error_thres
            self.model_buffer = ModelBuffer(alg_args.model_buffer_size)
            self.model_update_length = alg_args.model_update_length
            self.model_validate_interval = alg_args.model_validate_interval
            self.model_length_schedule = alg_args.model_length_schedule
            self.model_prob = alg_args.model_prob
        self.s, self.episode_len, self.episode_reward = self.env_learn.reset(), 0, 0

        
        # load pretrained model
        self.load_pretrained_model = alg_args.load_pretrained_model
        if self.model_based and self.load_pretrained_model:
            self.agent.load_model(alg_args.pretrained_model)
        
        if self.env_name == 'Real_Power':
            self.real_power_action_meam = (np.array([self.env_test.action_space.low]*self.env_test.n_agents) + np.array([self.env_test.action_space.high]*self.env_test.n_agents))/2
            self.real_power_action_var = (np.array([self.env_test.action_space.high]*self.env_test.n_agents) - np.array([self.env_test.action_space.low]*self.env_test.n_agents))/2
            self.running_state = ZFilter((self.env_learn.n_agents,self.env_learn.obs_size), clip=5.0)  #1.0
       
    def run(self):
        if self.model_based and not self.load_pretrained_model:
            for _ in trange(self.n_warmup):
                trajs = self.rollout_env()
                self.model_buffer.storeTrajs(trajs)
            self.updateModel(self.n_model_update_warmup) # Sample trajectories, then shorten them.

        for iter in trange(self.n_iter):
            # if iter % self.test_interval == 0:
            if iter % 10 == 0:
                mean_return = self.test(iter)


            if iter % 100 == 0:
                self.agent.save_nets(f'./checkpoints/{self.name}',iter)            

            trajs = self.rollout_env()  #  TO cheak: rollout n_step, maybe multi trajs
            t1=time.time()              
            if self.model_based:
                self.model_buffer.storeTrajs(trajs)
                # train the environment model
                if iter % 10 == 0:
                    self.updateModel()
            t2=time.time()
            print('t=',t2-t1)
                         
            agentInfo = []
            real_trajs = trajs
            for inner in trange(self.n_inner_iter):
                if self.model_based:
                    ## Use the model with a certain probability                  
                    use_model = np.random.uniform() < self.model_prob
                    if use_model:
                        if self.model_length_schedule is not None:
                            trajs = self.rollout_model(real_trajs, self.model_length_schedule(iter))
                        else:
                            trajs = self.rollout_model(real_trajs)
                    else:
                        trajs = trajs
                if self.clip_scheme is not None:
                    info = self.agent.updateAgent(trajs, self.clip_scheme(iter))     #  TO cheak: updata
                else:
                    info = self.agent.updateAgent(trajs)
                agentInfo.append(info)
                if self.agent.checkConverged(agentInfo):
                    break
            self.logger.log(inner_iter = inner + 1, iter=iter)

    def test(self,nnn):
        """
        The environment should return sth like [n_agent, dim] or [batch_size, n_agent, dim] in either numpy or torch.
        """
        time_t = time.time()
        length = self.test_length
        returns = []
        scaled = []
        lengths = []
        episodes = []
        S = []
        self.n_test = 20   #20
        record_each_step = True
        Ep = 0
        
        CR_rate = []
        Episode = []
        

        for i in trange(self.n_test):
            Episode.append(i)
            Ep += 1
            episode = []
            Step = []
            CR = []
            QL = []
            TL = []
            R = 0
            env = self.env_test    
            
            if self.env_name == 'eight' or self.env_name == 'ring':
                if i==0 and nnn == 0:
                    env.reset()    #for figure eight env
            else:       
                  
                env.reset()     # for another env
                #env.render()

            d, ep_ret, ep_len = np.array([False]), 0, 0
            #self.algo_name = "DMPO"
            #self.algo_name = "CPPO"
            

            
            while not(d.any() or (ep_len == length)):
                Step.append(ep_len)

                s = env.get_state_()
                #print("ssss=",s)
                
                        
                if self.env_name == 'PowerGrid' and env.n_agent==40:
                    s = self.running_state(s)

                elif self.env_name == 'Real_Power':
                    s = np.array(s)
                    s = self.running_state(s)



                
                s = torch.as_tensor(s, dtype=torch.float, device=self.device)
                
                

                
                a = self.agent.act(s,if_test=True).sample() # a is a tensor            
                a = a.detach().cpu().numpy() # might not be squeezed at the last dimension. env should deal with this though.
                



                if self.env_name == 'Real_Power':

                    #env.render()
                    if ep_len == 0:
                        time.sleep(5)  


                    

                    r, d, info = env.step(a)
                    s1 = env.get_state_()
                    s1 = np.array(s1)
                    s1 = self.running_state(s1)
                    r = np.array([r/env.n_agents]*env.n_agents)
                    #r = np.clip(r, -1, 0)
                    r = np.array([info["totally_controllable_ratio"]]*env.n_agents)
                    R += info["totally_controllable_ratio"]
                    
                    #print("ep_len=",ep_len)
                    #print("totally_controllable_ratio=",info["totally_controllable_ratio"])
                    #print("q_loss=",info["q_loss"])
                    #print("total_line_loss=",info["total_line_loss"])

                    CR.append(info["totally_controllable_ratio"])
                    QL.append(info["q_loss"])
                    TL.append(info["total_line_loss"])

                    d = [d]*env.n_agents




                else:    
                    s1, r, d, _ = env.step(a)
                    #print("a=",a)
                    #print("s1=",s1)
                    #print("r=",r)
                    #print("d=",d)

                    
                    
                episode += [(s.tolist(), a.tolist(), r.tolist())]
                d = np.array(d)
                ep_ret += r.sum()
                ep_len += 1
                #self.logger.log(interaction=None)
            
            CR_rate.append(R/239)
            from openpyxl import Workbook
            
            
            if record_each_step:
                wb_2 = Workbook() #´´½¨¹¤×÷²¾
                ws_2 = wb_2.active #¼¤»î¹¤×÷±í
                ws_2['A1'] = 'Step'
                ws_2['B1'] = 'totally_controllable_ratio'                            
                for i in range(len(Step)):                
                    ws_2.append([Step[i],dp(CR)[i]])              
                wb_2.save('./Real_Power_data/'+self.algo_name+'_'+str(Ep)+'_totally_controllable_ratio.xlsx')  
            
                wb_2 = Workbook() #´´½¨¹¤×÷²¾
                ws_2 = wb_2.active #¼¤»î¹¤×÷±í
                ws_2['A1'] = 'Step'
                ws_2['B1'] = 'q_loss'                            
                for i in range(len(Step)):                
                    ws_2.append([Step[i],dp(QL)[i]])              
                wb_2.save('./Real_Power_data/'+self.algo_name+'_'+str(Ep)+'_q_loss.xlsx')  
    
                wb_2 = Workbook() #´´½¨¹¤×÷²¾
                ws_2 = wb_2.active #¼¤»î¹¤×÷±í
                ws_2['A1'] = 'Step'
                ws_2['B1'] = 'total_line_loss'                            
                for i in range(len(Step)):                
                    ws_2.append([Step[i],dp(TL)[i]])              
                wb_2.save('./Real_Power_data/'+self.algo_name+'_'+str(Ep)+'_total_line_loss.xlsx')  


            if hasattr(env, 'rescaleReward'):
                scaled += [ep_ret]
                ep_ret = env.rescaleReward(ep_ret, ep_len)
            returns += [ep_ret]
            lengths += [ep_len]
            episodes += [episode]
        returns = np.stack(returns, axis=0)
        lengths = np.stack(lengths, axis=0)

        wb_2 = Workbook() #´´½¨¹¤×÷²¾
        ws_2 = wb_2.active #¼¤»î¹¤×÷±í
        ws_2['A1'] = 'Episode'
        ws_2['B1'] = 'controllable_ratio'                            
        for i in range(len(Episode)):                
            ws_2.append([Episode[i],dp(CR_rate)[i]])              
        wb_2.save('./Real_Power_data/'+self.algo_name+'_'+'_controllable_ratio.xlsx')  


        

        
        #self.logger.log(test_episode_reward=returns, test_episode_len=lengths, test_round=None)
        print(returns)
        print(f"{self.n_test} episodes average accumulated reward: {returns.mean()}")

        #self.logger.log(test_time=time.time()-time_t)
        return returns.mean()

    def rollout_env(self, length = 0):
        """
        The environment should return sth like [n_agent, dim] or [batch_size, n_agent, dim] in either numpy or torch.
        """
        time_t = time.time()
        if length <= 0:
            length = self.rollout_length
        env = self.env_learn
        trajs = []
        traj = TrajectoryBuffer(device=self.device)
        start = time.time()
        
        if self.env_name == 'Real_Power':
            totally_controllable_ratio = 0
            
        for t in range(length):
        # d, ep_len = np.array([False]), 0
        # while not(d.any() or (ep_len == length)):
            # ep_len+=1

            s = env.get_state_()    

            if self.env_name == 'PowerGrid' and env.n_agent==40:
                s = self.running_state(s)
            elif self.env_name == 'Real_Power':
                s = np.array(s)
                s = self.running_state(s)

       
            s = torch.as_tensor(s, dtype=torch.float, device=self.device)
            dist = self.agent.act(s)
            a = dist.sample()
            logp = dist.log_prob(a)         
            a = a.detach().cpu().numpy()
            
                    

            if (self.env_name == 'Monaco' and self.algo_name == 'IC3Net') or (self.env_name == 'Grid' and self.algo_name == 'IC3Net'):
                s1, r, d, _ = env.step(np.squeeze(a))


            elif self.env_name == 'PowerGrid' :
                if self.algo_name == 'IA2C' or self.algo_name == 'IC3Net':
                    s1, r, d, _ = env.step(np.squeeze(a))
                    d = np.array([d]*env.n_agent) 
                else:
                    s1, r, d, _ = env.step(a)
                    d = np.array([d]*env.n_agent)  
                if env.n_agent==40:
                    s1 = self.running_state(s1)

            elif self.env_name == 'Real_Power':



                r, d, info = env.step(a)

                s1 = env.get_state_()
                s1 = np.array(s1)
                s1 = self.running_state(s1)
                r = np.array([r/env.n_agents]*env.n_agents)
                #r = np.clip(r, -1, 0)
                d = np.array([d]*env.n_agents)
                totally_controllable_ratio += info["totally_controllable_ratio"]
                r = np.array([info["totally_controllable_ratio"]]*env.n_agents)
                r = np.float32(r)

     
            else:    
                s1, r, d, _ = env.step(a)

                
            traj.store(s, a, r, s1, d, logp)
            episode_r = r
            if hasattr(env, '_comparable_reward'):
                episode_r = env._comparable_reward()
            if episode_r.ndim > 1:
                episode_r = episode_r.mean(axis=0)
            #if episode_r.ndim == 1:
                #episode_r = episode_r.sum()
            self.episode_reward += episode_r
            self.episode_len += 1
            #self.logger.log(interaction=None)
            if self.episode_len == self.max_episode_len:
                d = np.zeros(d.shape, dtype=np.float32)
            d = np.array(d)
            
            # Do some rescales for different environments
#-----------------------------------------------------------------------------------------  
            # #for CACC_env(catchup and slowdown)
            if self.env_name == 'catchup' or self.env_name == 'slowdown':  
                if self.env_name == 'catchup':
                    if self.episode_len == self.max_episode_len:                 #for catchup                       
                        self.logger.log(episode_reward=self.episode_reward.sum()/600, episode_len = self.episode_len, episode=None)
                        try:
                            _, self.episode_reward, self.episode_len = self.env_learn.reset(), 0, 0#TODO:catch up the error
                        except Exception as e:
                            print('reset error!:', e)
                            _, self.episode_reward, self.episode_len = self.env_learn.reset(), 0, 0  # TODO:catch up the error
                            if self.model_based == False:
                                trajs += traj.retrieve()
                                traj = TrajectoryBuffer(device=self.device)
                    if self.episode_len == self.max_episode_len:
                        if self.model_based:
                            trajs += traj.retrieve()
                            traj = TrajectoryBuffer(device=self.device)    
                            
                elif self.env_name == 'slowdown':                                      
                    if d.any() or (self.episode_len == self.max_episode_len):      #for slowdown   
                        self.logger.log(episode_reward=self.episode_reward.sum()/600, episode_len = self.episode_len, episode=None)
                        try:
                            _, self.episode_reward, self.episode_len = self.env_learn.reset(), 0, 0#TODO:catch up the error
                        except Exception as e:
                            print('reset error!:', e)
                            _, self.episode_reward, self.episode_len = self.env_learn.reset(), 0, 0  # TODO:catch up the error
                            if self.model_based == False:
                                trajs += traj.retrieve()
                                traj = TrajectoryBuffer(device=self.device)                           
                    if self.episode_len == self.max_episode_len:
                        if self.model_based:
                            trajs += traj.retrieve()
                            traj = TrajectoryBuffer(device=self.device)
#----------------------------------------------------------------------------------------- 
            elif self.env_name == 'eight' or self.env_name == 'ring':          
                # if d.any() or (self.episode_len == self.max_episode_len):     
                if self.episode_len == self.max_episode_len:                                
                    self.logger.log(episode_reward=self.episode_reward.sum(), episode_len = self.episode_len, episode=None)
                    try:
                        self.episode_reward, self.episode_len = 0, 0#TODO:catch up the error
                    except Exception as e:
                        print('reset error!:', e)
                        self.episode_reward, self.episode_len =  0, 0  # TODO:catch up the error
                        if self.model_based == False:
                            trajs += traj.retrieve()
                            traj = TrajectoryBuffer(device=self.device)
                if self.episode_len == self.max_episode_len:
                    if self.model_based:
                        trajs += traj.retrieve()
                        traj = TrajectoryBuffer(device=self.device)



            elif self.env_name == 'PowerGrid':
            # for other_env
                if d.any() or (self.episode_len == self.max_episode_len):      
                # if self.episode_len == self.max_episode_len:                 
                    
                    self.logger.log(episode_reward=self.episode_reward.sum()/env.T, episode_len = self.episode_len, episode=None)
                    try:
                        _, self.episode_reward, self.episode_len = self.env_learn.reset(), 0, 0#TODO:catch up the error
                    except Exception as e:
                        print('reset error!:', e)
                        _, self.episode_reward, self.episode_len = self.env_learn.reset(), 0, 0  # TODO:catch up the error
                        if self.model_based == False:
                            trajs += traj.retrieve()
                            traj = TrajectoryBuffer(device=self.device)
                            
                if self.episode_len == self.max_episode_len:
                    if self.model_based:
                        trajs += traj.retrieve()
                        traj = TrajectoryBuffer(device=self.device)

            elif self.env_name == 'Real_Power':
            # for other_env
                if d.any() or (self.episode_len == self.max_episode_len):      
                # if self.episode_len == self.max_episode_len:                 
                    
                    #self.logger.log(episode_reward=self.episode_reward.sum(), episode_len = self.episode_len, episode=None, totally_controllable_ratio=totally_controllable_ratio)
                    try:
                        _, self.episode_reward, self.episode_len = self.env_learn.reset(), 0, 0#TODO:catch up the error
                    except Exception as e:
                        print('reset error!:', e)
                        _, self.episode_reward, self.episode_len = self.env_learn.reset(), 0, 0  # TODO:catch up the error
                        if self.model_based == False:
                            trajs += traj.retrieve()
                            traj = TrajectoryBuffer(device=self.device)
                            
                if self.episode_len == self.max_episode_len:
                    if self.model_based:
                        trajs += traj.retrieve()
                        traj = TrajectoryBuffer(device=self.device)


            else:
            # for other_env
                if d.any() or (self.episode_len == self.max_episode_len):      
                # if self.episode_len == self.max_episode_len:                 
                    
                    self.logger.log(episode_reward=self.episode_reward.sum(), episode_len = self.episode_len, episode=None)
                    try:
                        _, self.episode_reward, self.episode_len = self.env_learn.reset(), 0, 0#TODO:catch up the error
                    except Exception as e:
                        print('reset error!:', e)
                        _, self.episode_reward, self.episode_len = self.env_learn.reset(), 0, 0  # TODO:catch up the error
                        if self.model_based == False:
                            trajs += traj.retrieve()
                            traj = TrajectoryBuffer(device=self.device)
                            
                if self.episode_len == self.max_episode_len:
                    if self.model_based:
                        trajs += traj.retrieve()
                        traj = TrajectoryBuffer(device=self.device)
#--------------------------------------------------------------------------------------    
        end = time.time()
        print('time in 1 episode is ',end-start)
        trajs += traj.retrieve(length=self.max_episode_len)
        #self.logger.log(env_rollout_time=time.time()-time_t)
        return trajs
    
    # Use the environment model to collect data
    def rollout_model(self, trajs, length=0):
        time_t = time.time()
        n_traj = self.n_traj
        if length <= 0:
            length = self.model_traj_length
        s = [traj['s'] for traj in trajs]

        s = torch.stack(s, dim=0)
        b, T, n, depth = s.shape
        s = s.view(-1, n, depth)
        idxs = torch.randint(low=0, high=b * T, size=(n_traj,), device=self.device)
        s = s.index_select(dim=0, index=idxs)

        trajs = TrajectoryBuffer(device=self.device)
        for _ in range(length):
            #a, logp = self.agent.act(s, requires_log=True)
            dist = self.agent.act(s)
            a = dist.sample()
            
            
            #if self.env_name == 'Real_Power':
                #a = translate_action_2(self.env_learn.args.action_bias, self.env_learn.args.action_scale, a)
            


            
            logp = dist.log_prob(a)
            r, s1, d, _ = self.agent.model_step(s, a)
            
            #if self.env_name == 'Real_Power':
                #r = torch.clamp(r, min=-1.0, max=0)
            


            if self.env_name == 'UAV_9d':
               env = self.env_learn
               s = env.get_model_state(s,self.device)
               s1 = env.get_model_state(s1,self.device)
               r = env.get_model_reward(s1,self.device)

            trajs.store(s, a, r, s1, d, logp)
            s = s1
        trajs = trajs.retrieve()
        self.logger.log(model_rollout_time=time.time()-time_t)
        return trajs
    

    def updateModel(self, n=0):
        if n <= 0:
            n = self.n_model_update
        for i_model_update in trange(n):
            trajs = self.model_buffer.sampleTrajs(self.model_batch_size)
            trajs = [traj.getFraction(length=self.model_update_length) for traj in trajs]
            
            self.agent.updateModel(trajs, length=self.model_update_length)

            if i_model_update % self.model_validate_interval == 0:
                validate_trajs = self.model_buffer.sampleTrajs(self.model_batch_size)
                validate_trajs = [traj.getFraction(length=self.model_update_length) for traj in validate_trajs]
                rel_error = self.agent.validateModel(validate_trajs, length=self.model_update_length)
                if rel_error < self.model_error_thres:
                    break
        self.logger.log(model_update = i_model_update + 1)

    def testModel(self, n = 0):
        trajs = self.model_buffer.sampleTrajs(self.model_batch_size)
        trajs = [traj.getFraction(length=self.model_update_length) for traj in trajs]
        return self.agent.validateModel(trajs, length=self.model_update_length)

